# -*- coding: utf-8 -*-
"""生成 OA 交换机巡检项阈值配置表"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = 'OA交换机巡检项-知识城'

# 样式
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
wrap = Alignment(wrap_text=True, vertical='top')
thin = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
data_font = Font(name='微软雅黑', size=10)
edit_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
new_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

headers = ['序号', '检查项', '命令', '适用设备类型', '判断', 'parser', 'checker',
           'parser_config', 'checker_config', '现网阈值/期望值（待填）', '备注']
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin

items = [
    (1, 'CPU利用率', 'display cpu-usage', '所有', 'B', 'regex', 'threshold',
     '{"pattern":"(\\d+)%","group":1,"cast":"float"}',
     '{"warning":80,"operator":"<"}',
     'CPU<80% 无持续升高', ''),
    (2, '内存利用率', 'display memory', '所有', 'B', 'regex', 'threshold',
     '{"pattern":"Mem:\\s+\\d+\\s+\\d+\\s+\\d+\\s+\\d+\\s+\\d+\\s+\\d+\\s+(\\d+\\.?\\d*)%","group":1,"cast":"float"}',
     '{"warning":20,"operator":">"}',
     'FreeRatio<20%为异常', '当前55%正常'),
    (3, '风扇', 'display fan', '所有', 'B', 'raw', 'custom',
     '', '{"func":"check_fan"}',
     '所有风扇Normal', '已有'),
    (4, '电源', 'display power', '所有', 'B', 'raw', 'custom',
     '', '{"func":"check_power"}',
     '所有电源Normal', '已有'),
    (5, '温度', 'display environment', '所有', 'B', 'raw', 'custom',
     '', '{"func":"check_env","temp_warning":60}',
     '温度<60度 无告警', '已有，当前22度'),
    (6, '单板', 'display device', '所有', 'B', 'raw', 'custom',
     '', '{"func":"check_device"}',
     '单板Normal', '已有'),
    (7, '接口DOWN数', 'display interface brief', '所有', 'B', 'raw', 'custom',
     '', '{"func":"check_ifbrief"}',
     '物理DOWN = 规划值', '已有'),
    (8, 'STP状态', 'display stp brief', '接入', 'B', 'raw', 'custom',
     '', '{"func":"check_stp","root_expected":"非根桥"}',
     '非根桥，DESI/FWD', '新 checker'),
    (9, 'VLAN清单', 'display vlan brief', '接入', 'C', 'raw', 'custom',
     '', '{"func":"check_vlan","expected_vlans":[1,100,101]}',
     'VLAN集合=期望', '新 checker'),
    (10, '链路聚合', 'display link-aggregation summary', '所有', 'B', 'raw', 'custom',
     '', '{"func":"check_agg"}',
     'Selected数=成员数', '已有'),
    (11, '日志缓冲', 'display logbuffer', '所有', 'A', 'strip_ts', 'baseline',
     '{"patterns":["%\\w{3}\\s+\\d+\\s+\\d{2}:\\d{2}:\\d{2}:\\d{3}\\s+\\d{4}"]}',
     '{"similarity":0.95}',
     '无新增Error/Critical', '时间戳特殊'),
    (12, '路由表', 'display ip routing-table', '接入', 'C', 'raw', 'custom',
     '', '{"func":"check_routing_table","expected_routes":["0.0.0.0/0 via 10.202.224.126"]}',
     '默认路由存在下一跳正确', '新 checker'),
    (13, 'ARP冲突', 'display arp user-ip-conflict record', '接入', 'A', 'raw', 'custom',
     '', '{"func":"check_arp"}',
     '冲突记录清零', '已有'),
    (14, 'NTP状态', 'display ntp status', '所有', 'B', 'raw', 'contains',
     '', '{"keyword":"synchronized"}',
     'Clock synchronized', ''),
    (15, '版本', 'display version', '所有', '-', 'raw', 'contains',
     '', '{}', '仅采集', ''),
    (16, '当前配置', 'display current-configuration', '所有', '-', 'raw', 'contains',
     '', '{}', '仅采集备份', ''),
    (17, '稳定状态', 'display system stable state', '所有', '-', 'raw', 'contains',
     '', '{}', '仅采集', ''),
    (18, 'LLDP邻居', 'display lldp neighbor-information list', '所有', '-', 'raw', 'contains',
     '', '{}', '仅采集', ''),
    (19, '光模块诊断', 'display transceiver diagnosis interface', '所有', '-', 'raw', 'contains',
     '', '{}', '仅采集', ''),
    (20, '错包入向', 'display counters inbound interface', '所有', '-', 'raw', 'contains',
     '', '{}', '仅采集', ''),
    (21, '错包出向', 'display counters outbound interface', '所有', '-', 'raw', 'contains',
     '', '{}', '仅采集', ''),
    (22, 'Flash空间', 'dir flash:/', '所有', '-', 'raw', 'contains',
     '', '{}', '仅采集', ''),
]

for i, row_data in enumerate(items):
    row = i + 2
    for col, val in enumerate(row_data, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = data_font
        cell.alignment = wrap
        cell.border = thin
    ws.cell(row=row, column=10).fill = edit_fill
    if '新 checker' in str(row_data[10]):
        ws.cell(row=row, column=8).fill = new_fill
        ws.cell(row=row, column=9).fill = new_fill

widths = [5, 18, 32, 12, 5, 10, 12, 45, 45, 32, 22]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = 'A2'

output = '巡检项阈值配置表_OA交换机_知识城.xlsx'
wb.save(output)
print(f'OK: {output} generated (22 items)')
print('   红色 = 新 checker  |  黄色 = 待填阈值')