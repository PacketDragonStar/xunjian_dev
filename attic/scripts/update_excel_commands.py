# -*- coding: utf-8 -*-
"""
按角色填充两个设备Excel的"命令(覆盖)"列。
基于: 巡检命令角色模板_知识城_2026-07-15.xlsx
"""
import os
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter

# 文件路径
ROLE_TEMPLATE = r'C:\Users\ZSS\Desktop\化龙\化龙\化龙配置\巡检命令角色模板_知识城_2026-07-15.xlsx'
HL_EXCEL = r'C:\Users\ZSS\Desktop\化龙\化龙\化龙配置\network_inspection\化龙设备清单_带巡检命令.xlsx'
ZSC_EXCEL = r'C:\Users\ZSS\Desktop\化龙\化龙\化龙配置\network_inspection\知识城设备清单_带巡检命令.xlsx'

# ── 1. 读角色模板，获取角色→命令列表 ──
print('📖 读取角色模板...')
xl = pd.ExcelFile(ROLE_TEMPLATE)
df_role = pd.read_excel(xl, '角色巡检模板')
df_judge = pd.read_excel(xl, '巡检命令判断明细(按角色)')

# 从角色模板提取命令
role_cmds = {}
for _, row in df_role.iterrows():
    role = row['角色']
    cmds_text = row['标准巡检命令(逐条执行, 编号)']
    cmds = [c.strip() for c in cmds_text.split('\n') if c.strip()]
    # 去掉编号前缀 "1. " / "2. " 等
    cmds = [c.split('. ', 1)[1] if '. ' in c[:5] else c for c in cmds]
    role_cmds[role] = cmds
    print(f'  {role}: {len(cmds)} 条命令')

# ── 2. 从判断明细补充角色→命令映射（更精确） ──
print('\n📖 读取判断明细...')
from collections import defaultdict
role_cmds_detail = defaultdict(list)
for _, row in df_judge.iterrows():
    cmd = row['完整命令']
    roles_text = str(row['适用角色'])
    if '全部' in roles_text or '所有' in roles_text:
        for r in role_cmds:
            role_cmds_detail[r].append(cmd)
    elif '防火墙' in roles_text:
        role_cmds_detail['防火墙'].append(cmd)
    elif '核心' in roles_text:
        role_cmds_detail['核心交换机'].append(cmd)
    elif '接入' in roles_text or '汇聚' in roles_text:
        role_cmds_detail['接入交换机'].append(cmd)
    elif 'OA' in roles_text:
        role_cmds_detail['OA交换机'].append(cmd)
    elif '存储' in roles_text:
        role_cmds_detail['存储交换机'].append(cmd)
    elif 'IDC' in roles_text:
        role_cmds_detail['IDC交换机'].append(cmd)

# 补充命令: 从之前的 commands_inspection.yaml
extra_common = [
    'display version', 'display current-configuration', 'display ntp status',
    'display counters inbound interface', 'display counters outbound interface',
    'display lldp neighbor-information list', 'display transceiver diagnosis interface',
    'display system stable state', 'dir flash:/'
]

# ── 3. 设备名 → 角色 映射函数 ──
def role_of(name: str) -> str:
    s = (name or '').lower()
    if s.startswith('fw'):   return '防火墙'
    if s.startswith('csw'):  return '核心交换机'
    if s.startswith('srp'):  return '路由器'
    if s.startswith('oas'):  return 'OA交换机'
    if s.startswith('psw'):  return '存储交换机'
    if s.startswith('idc'):  return 'IDC交换机'
    if s.startswith(('asw', 'usw')): return '接入交换机'
    if s.startswith(('dci', 'dsw')):  return '接入交换机'  # DCI/汇聚作为接入处理
    return '接入交换机'

# ── 4. 为每个角色拼接最终命令（含命令序号） ──
def build_commands_for_role(role: str) -> str:
    """返回 Excel 单元格格式的命令字符串（每条一行，Alt+Enter 换行）"""
    cmds = list(role_cmds.get(role, []))
    # 去重后加入通用补充命令
    seen = set(cmd.lower() for cmd in cmds)
    for ec in extra_common:
        if ec.lower() not in seen:
            cmds.append(ec)
    # 纯命令列表，每行一条（Excel 单元格内 Alt+Enter 换行）
    return '\n'.join(cmds)

# ── 5. 更新两个 Excel ──
def update_excel(filepath: str, label: str, output_path: str = None):
    if output_path is None:
        base, ext = os.path.splitext(filepath)
        output_path = f'{base}_已填充命令{ext}'
    print(f'\n🔄 更新 {label}: {filepath} → {output_path}')
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active

    # 找"命令(覆盖)"列
    headers = [str(c.value or '').strip() for c in ws[1]]
    cmd_col = None
    name_col = None
    for i, h in enumerate(headers):
        if '命令' in h:
            cmd_col = i
        if '设备名称' in h or '设备名' in h:
            name_col = i

    if cmd_col is None:
        print(f'  ❌ 未找到"命令"列')
        return
    if name_col is None:
        print(f'  ❌ 未找到"设备名称"列')
        return

    cmd_col_letter = get_column_letter(cmd_col + 1)
    name_col_letter = get_column_letter(name_col + 1)

    updated = 0
    skipped = 0
    for row in range(2, ws.max_row + 1):
        name = str(ws.cell(row=row, column=name_col + 1).value or '').strip()
        if not name:
            continue
        role = role_of(name)
        if role == '路由器':
            skipped += 1
            continue
        cmds_text = build_commands_for_role(role)
        ws.cell(row=row, column=cmd_col + 1).value = cmds_text
        updated += 1
        print(f'  {name} → {role} ({cmds_text.count(chr(10))+1} 条)')

    # 调整列宽
    ws.column_dimensions[cmd_col_letter].width = 60

    wb.save(output_path)
    wb.close()
    print(f'  ✅ 更新 {updated} 台设备, 跳过 {skipped} 台(路由器)')
    print(f'  📁 保存至: {output_path}')

# ── 6. 执行 ──
print('\n' + '='*60)
print('各角色最终命令一览:')
for role in role_cmds:
    cmds = build_commands_for_role(role)
    count = cmds.count('\n') + 1
    print(f'  {role}: {count} 条')

# 输出到项目目录（避免文件锁问题）
OUT_DIR = os.path.dirname(os.path.abspath(__file__))
update_excel(HL_EXCEL, '化龙设备清单_带巡检命令.xlsx',
             os.path.join(OUT_DIR, '化龙设备清单_已填充命令.xlsx'))
update_excel(ZSC_EXCEL, '知识城设备清单_带巡检命令.xlsx',
             os.path.join(OUT_DIR, '知识城设备清单_已填充命令.xlsx'))

print('\n✅ 完成! 两个 Excel 的"命令(覆盖)"列已按角色填充。')
print('   现在执行: python inspection.py --excel 化龙设备清单_带巡检命令.xlsx --commands commands_inspection.yaml')
print('   命令将从 Excel 读取（优先级最高），不再依赖 YAML 的通用组。')