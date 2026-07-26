你# -*- coding: utf-8 -*-
"""生成巡检项阈值配置表 Excel"""
import sys
sys.stdout.reconfigure(encoding='utf-8')

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = Workbook()

# ============================================================
# 通用样式
# ============================================================
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
wrap = Alignment(wrap_text=True, vertical='top')
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
data_font = Font(name='微软雅黑', size=10)
edit_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

def write_header(ws, headers, row=1):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

def write_row(ws, row, values):
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = data_font
        cell.alignment = wrap
        cell.border = thin_border

def set_col_widths(ws, widths):
    from openpyxl.utils import get_column_letter
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

# ============================================================
# Sheet 1: 巡检项阈值配置表
# ============================================================
ws1 = wb.active
ws1.title = '巡检项阈值配置表'

headers1 = ['序号', '检查项', '命令', '适用设备类型', '判断方式',
            'parser', 'checker', 'parser_config', 'checker_config',
            'Excel原始阈值', '✏️ 现网阈值（待填）', '备注']
write_header(ws1, headers1)

data1 = [
    # 一、所有设备通用 (rows 2-9)
    [1, 'CPU利用率', 'display cpu-usage', '所有设备', 'B', 'regex', 'threshold',
     '{"pattern": "([\\d\\.]+)%", "group": 1, "cast": "float"}',
     '{"warning": 80, "operator": "<"}',
     'CPU利用率<80%，无持续上升趋势；多核需逐核查看', '', ''],
    [2, '内存利用率', 'display memory', '所有设备', 'B', 'regex', 'threshold',
     '{"pattern": "([\\d\\.]+)%", "group": 1, "cast": "float"}',
     '{"warning": 85, "operator": "<"}',
     '内存利用率<85%，无持续增长', '', ''],
    [3, '风扇状态', 'display fan', '所有设备', 'B', 'raw', 'custom',
     None, '{"func": "check_fan"}',
     '所有风扇 State 为 Normal/Work', '', '✅ 已有，无需新写'],
    [4, '电源状态', 'display power', '所有设备', 'B', 'raw', 'custom',
     None, '{"func": "check_power"}',
     '所有电源 Status 为 Normal/Supply', '', '✅ 已有'],
    [5, '温度环境', 'display environment', '所有设备', 'B', 'raw', 'custom',
     None, '{"func": "check_env", "temp_warning": 60}',
     '温度/电压在阈值内，无温度告警', '', '修改 temp_warning 为现网阈值'],
    [6, '单板状态', 'display device', '所有设备', 'B', 'raw', 'custom',
     None, '{"func": "check_device"}',
     '所有单板 Status 为 Normal，无 Fault', '', '✅ 已有'],
    [7, '接口DOWN数（规则）', 'display interface brief', '所有设备', 'B', 'raw', 'custom',
     None, '{"func": "check_ifbrief"}',
     '物理DOWN接口数 ≤ 规划值（per-device extra.down_ok）', '', '✅ 已有；down_ok 在设备 extra 字段设'],
    [8, '接口CRC/错包（基线）', 'display interface brief', '所有设备', 'A', 'raw', 'baseline',
     None, '{"similarity": 1.0}',
     '对比上次基线CRC/错包增量应为0', '', '#7拆为两个CheckItem：7A规则+7B基线'],

    # 二、日志 (row 10)
    [9, '日志缓冲', 'display logbuffer', '所有设备', 'A', 'strip_ts', 'baseline',
     '{"patterns": ["\\d{4}-\\d{2}-\\d{2}\\s+\\d{2}:\\d{2}:\\d{2}"]}',
     '{"similarity": 0.95}',
     '无新增 Error/Critical/告警级日志；有则逐条排查', '', '去时间戳后对比，避免时间变化导致全量diff'],

    # 三、防火墙专用 (rows 11-16)
    [10, '安全域成员', 'display zone', '防火墙', 'C', 'raw', 'custom',
     None, '{"func": "check_zone", "expected": {"OutBand": [], "InBand": [], "GLQ_monitor_MGMT": []}}',
     '各安全域 import 接口集合与 running-config 一致', '', '⚠️ 需按现网补填各域接口列表'],
    [11, '安全策略命中', 'display security-policy statistics', '防火墙', 'A', 'raw', 'baseline',
     None, '{"similarity": 1.0}',
     '对比基线各 rule 命中计数，异常增长/归零需排查', '', ''],
    [12, '安全策略规则核查', 'display security-policy ip rule all', '防火墙', 'C', 'raw', 'custom',
     None, '{"func": "check_security_policy_zone"}',
     '每条规则的dest_zone须等于目的IP真实出域', '', '复杂逻辑，需建zone→接口映射表'],
    [13, '双机热备状态', 'display rbm', '防火墙', 'B', 'raw', 'custom',
     None, '{"func": "check_rbm"}',
     '主墙 RBM_P 为 Active，备墙 Inactive', '', '需新增 checker'],
    [14, '防火墙-路由表', 'display ip routing-table', '防火墙', 'C', 'raw', 'custom',
     None, '{"func": "check_routing_table", "expected_routes": ["0.0.0.0/0 via 请填下一跳IP"]}',
     '默认路由/静态路由存在且下一跳正确', '', '⚠️ 需补填期望路由下一跳IP'],
    [15, '会话表', 'display session table', '防火墙', 'B', 'regex', 'custom',
     '{"pattern": "([\\d]+)", "group": 1, "cast": "int"}',
     '{"func": "check_session", "max_sessions": 500000}',
     '并发会话数在正常范围，无异常暴增', '', '⚠️ 需修改 max_sessions 为现网上限'],

    # 四、化龙核心交换机专用 (rows 17-22)
    [16, 'OSPF邻居（化龙）', 'display ospf peer', '化龙核心交换机', 'B', 'raw', 'custom',
     None, '{"func": "check_ospf_peer", "expected_full_count": 0}',
     '所有邻居 State=Full，数量==配置 area 内邻居数', '', '⚠️ 需填写期望 Full 邻居数'],
    [17, 'BGP邻居（化龙）', 'display bgp peer', '化龙核心交换机', 'B', 'raw', 'custom',
     None, '{"func": "check_bgp_peer", "expected_established": 1}',
     'Established 数量==配置 peer 数（化龙1个，peer 10.202.136.170 as 65533）', '', '⚠️ 确认期望 Established 数'],
    [18, 'VRRP状态（化龙）', 'display vrrp', '化龙核心交换机', 'B', 'raw', 'custom',
     None, '{"func": "check_vrrp"}',
     'Master/Backup 角色符合规划', '', '✅ 已有；per-device extra.vrrp_master'],
    [19, 'STP状态（化龙）', 'display stp brief', '化龙核心交换机', 'B', 'raw', 'custom',
     None, '{"func": "check_stp", "root_expected": "本端"}',
     '实例根桥为本端（化龙 csw 为 root primary）', '', '需新增 checker'],
    [20, 'VLAN清单（化龙）', 'display vlan brief', '化龙核心交换机', 'C', 'raw', 'custom',
     None, '{"func": "check_vlan", "expected_vlans": []}',
     'VLAN 集合==配置允许 VLAN', '', '⚠️ 需补填化龙核心 VLAN 列表'],
    [21, '链路聚合（化龙）', 'display link-aggregation summary', '化龙核心交换机', 'B', 'raw', 'custom',
     None, '{"func": "check_agg"}',
     'Selected 端口数==配置成员数（BA1-6/21-32/34/36/100）', '', '✅ 已有；per-device extra'],

    # 五、知识城核心交换机专用 (rows 23-29)
    [22, 'OSPF多实例邻居（知识城）', 'display ospf vpn-instance EIP_MGMT peer', '知识城核心交换机', 'B', 'raw', 'custom',
     None, '{"func": "check_ospf_peer", "instances": {"EIP_MGMT": 0, "Internet": 0, "EIP_YW": 0, "Wocloud_ZWWW": 0, "ECM_YW": 0, "Wocloud_HLW": 0}}',
     '各实例邻居 Full 数==配置（area 0.0.0.0）', '', '⚠️ 需填写每个实例期望的 Full 邻居数'],
    [23, 'VRRP vrid203', 'display vrrp vrid 203', '知识城核心交换机', 'B', 'raw', 'custom',
     None, '{"func": "check_vrrp"}',
     'Master 角色正确，VIP 100.64.0.254 可通（pri 200）', '', '✅ 已有'],
    [24, 'M-LAG状态', 'display m-lag summary', '知识城核心交换机', 'B', 'raw', 'custom',
     None, '{"func": "check_mlag"}',
     '状态 Active，MAD 口正常，无 MAD 冲突', '', '需新增 checker'],
    [25, 'STP状态（知识城）', 'display stp brief', '知识城核心交换机', 'B', 'raw', 'custom',
     None, '{"func": "check_stp", "root_expected": "按规划"}',
     '根桥/端口角色符合规划', '', '⚠️ 确认根桥是哪个设备'],
    [26, 'VLAN清单（知识城核心）', 'display vlan brief', '知识城核心交换机', 'C', 'raw', 'custom',
     None, '{"func": "check_vlan", "expected_vlans": [100, 172, 173, 174, 175, 176, 177, 178, 179, 180, 181, 182, 183, 184, 185, 186, 187, 188, 189, 190, 191, 192, 193, 194, 195, 196, 197, 198, 199, 200, 201, 202, 203, 204, 380, 381, 2000, 2001, 2002, 2003, 2004, 2005, 2006, 2007, 2008, 2009, 2010, 2011, 2012, 2013, 2014, 2015, 2016, 2017, 2018, 2019, 2020, 2021, 2022, 2023, 2024, 2025, 2026, 2027, 2028, 2029, 2030, 2031, 2032, 2033, 2034, 2035, 2036, 2037, 2038, 2039, 2040, 2041, 2042, 2043, 2044, 2045, 2046, 2047, 2048, 2049, 2050, 2051, 2052, 2053, 2054, 2055, 2056, 2057, 2058, 2059, 2060]}',
     'VLAN 100/172-204/380-381/2000-2600', '', '⚠️ 需确认最终 VLAN 列表'],
    [27, '链路聚合（知识城）', 'display link-aggregation summary', '知识城核心交换机', 'B', 'raw', 'custom',
     None, '{"func": "check_agg"}',
     'Selected 数==成员数', '', '✅ 已有'],
    [28, 'NQA/Track状态', 'display track', '知识城核心交换机', 'B', 'raw', 'custom',
     None, '{"func": "check_track", "expected_tracks": 12}',
     'Track 1-12 状态均为 Positive（to_ali/to_ecm_yw/to_eip_yw/to_int_hw_sp）', '', '⚠️ 确认Track数量和名称'],

    # 六、接入/汇聚/OA/存储/IDC/上行交换机 (rows 30-34)
    [29, 'STP状态（非核心）', 'display stp brief', '接入/汇聚/OA/存储/IDC/上行交换机', 'B', 'raw', 'custom',
     None, '{"func": "check_stp", "root_expected": "非根桥"}',
     '非根桥，access/trunk 端口角色正确', '', ''],
    [30, 'VLAN清单（非核心）', 'display vlan brief', '接入/汇聚/OA/存储/IDC/上行交换机', 'C', 'raw', 'custom',
     None, '{"func": "check_vlan", "expected_vlans": []}',
     '允许 VLAN==本设备 permit VLAN', '', '⚠️ 需补填各组设备的 permit VLAN 列表'],
    [31, '链路聚合（非核心）', 'display link-aggregation summary', '接入/汇聚/OA/存储/IDC/上行交换机', 'B', 'raw', 'custom',
     None, '{"func": "check_agg"}',
     'Selected 端口数==配置成员数', '', '✅ 已有'],
    [32, '路由表（非核心）', 'display ip routing-table', '接入/汇聚/OA/存储/IDC/上行交换机', 'C', 'raw', 'custom',
     None, '{"func": "check_routing_table", "expected_routes": ["请填默认路由"]}',
     '默认路由/静态路由存在且下一跳正确', '', '⚠️ 需补填期望路由'],
    [33, 'ARP冲突记录', 'display arp user-ip-conflict', '知识城接入/汇聚/OA/存储/IDC/上行', 'A', 'raw', 'custom',
     None, '{"func": "check_arp"}',
     '冲突记录应清零（知识城ARP告警根因）；非零需排查', '', '✅ 已有'],
]

# 分类标题定义
categories = {
    2:  '一、所有设备通用（#1-#8）',
    10: '二、日志（#9）',
    11: '三、防火墙专用（#10-#15）',
    17: '四、化龙核心交换机专用（#16-#21）',
    23: '五、知识城核心交换机专用（#22-#28）',
    30: '六、接入/汇聚/OA/存储/IDC/上行交换机（#29-#33）',
}

current_row = 2
for i, row_data in enumerate(data1):
    row_num = i + 2
    if row_num in categories:
        # 写入分类标题行
        ws1.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=12)
        cell = ws1.cell(row=current_row, column=1, value=categories[row_num])
        cell.font = Font(name='微软雅黑', size=11, bold=True, color='333333')
        cell.alignment = Alignment(horizontal='left', vertical='center')
        for col in range(1, 13):
            ws1.cell(row=current_row, column=col).border = thin_border
        current_row += 1

    write_row(ws1, current_row, row_data)
    # 第11列（现网阈值）高亮黄色
    ws1.cell(row=current_row, column=11).fill = edit_fill
    current_row += 1

set_col_widths(ws1, [5, 22, 30, 20, 8, 12, 12, 40, 50, 40, 30, 30])
ws1.freeze_panes = 'A2'

# ============================================================
# Sheet 2: 需新增Checker清单
# ============================================================
ws2 = wb.create_sheet('需新增Checker清单')

headers2 = ['序号', 'checker名称', '类型', '对应检查项', '检查方式', '输入参数', '判断逻辑', '优先级']
write_header(ws2, headers2)

checkers2 = [
    [1, 'check_cpu', 'B', 'CPU利用率', '阈值判断', 'parsed(正则提取数值), cfg.warning, cfg.operator', 'CPU利用率 < warning（默认80%），多核逐核检查', 'P0'],
    [2, 'check_memory', 'B', '内存利用率', '阈值判断', 'parsed(正则提取数值), cfg.warning, cfg.operator', '内存利用率 < warning（默认85%）', 'P0'],
    [3, 'check_stp', 'B', 'STP状态（核心+接入）', '规则判断', 'parsed(原始文本), cfg.root_expected("本端"/"非根桥"/"按规划")', '根桥/端口角色符合规划；非根桥设备不应为根', 'P0'],
    [4, 'check_ospf_peer', 'B', 'OSPF邻居（含多实例）', '规则判断', 'parsed(原始文本), cfg.expected_full_count 或 cfg.instances', '所有邻居 State=Full 且数量==期望值', 'P0'],
    [5, 'check_bgp_peer', 'B', 'BGP邻居', '规则判断', 'parsed(原始文本), cfg.expected_established', 'Established 邻居数==期望值', 'P0'],
    [6, 'check_rbm', 'B', '双机热备状态', '规则判断', 'parsed(原始文本)', '主墙 RBM_P=Active，备墙=Inactive', 'P0'],
    [7, 'check_mlag', 'B', 'M-LAG状态', '规则判断', 'parsed(原始文本)', '状态 Active，MAD口正常，无MAD冲突', 'P0'],
    [8, 'check_track', 'B', 'NQA/Track状态', '规则判断', 'parsed(原始文本), cfg.expected_tracks', 'Track 1-N 全 Positive', 'P0'],
    [9, 'check_session', 'B', '会话表', '阈值判断', 'parsed(正则提取数值), cfg.max_sessions', '并发会话数 < max_sessions', 'P1'],
    [10, 'check_vlan', 'C', 'VLAN清单', '配置一致性', 'parsed(提取VLAN ID集合), cfg.expected_vlans', '实际VLAN集合 == 期望VLAN集合（双向比对：缺少+多了都报）', 'P0'],
    [11, 'check_zone', 'C', '安全域成员', '配置一致性', 'parsed(提取zone→接口映射), cfg.expected', '各zone的import接口集合 == 期望值', 'P1'],
    [12, 'check_routing_table', 'C', '路由表', '配置一致性', 'parsed(提取路由条目), cfg.expected_routes', '期望路由全部存在且下一跳正确', 'P0'],
    [13, 'check_security_policy_zone', 'C', '安全策略规则zone', '配置一致性', 'parsed(提取rule→dest_zone映射)', '每条rule的dest_zone==路由出接口对应zone', 'P1'],
    [14, 'check_logbuffer', 'B', '日志缓冲Error计数', '规则判断', 'parsed(原始文本)', '统计 Error/Critical 级别日志条数（辅助baseline判断）', 'P1'],
]

for i, row_data in enumerate(checkers2):
    write_row(ws2, i+2, row_data)

set_col_widths(ws2, [5, 25, 6, 25, 12, 40, 50, 6])
ws2.freeze_panes = 'A2'

# ============================================================
# Sheet 3: 设备分组与CheckItem绑定关系
# ============================================================
ws3 = wb.create_sheet('设备分组与CheckItem绑定')

headers3 = ['分组名称', '站点', '适用角色', '绑定巡检项序号范围', '绑定巡检项名称说明']
write_header(ws3, headers3)

groups3 = [
    ['化龙-防火墙', '化龙', '防火墙', '1-15', '通用1-9 + 防火墙专用10-15（安全域/策略命中/策略规则/RBM/路由表/会话表）'],
    ['化龙-核心交换机', '化龙', '核心交换机', '1-9, 16-21', '通用1-9 + 化龙核心专用16-21（OSPF/BGP/VRRP/STP/VLAN/链路聚合）'],
    ['化龙-接入交换机', '化龙', '接入交换机', '1-9, 29-33', '通用1-9 + 非核心通用29-33（STP/VLAN/链路聚合/路由表/ARP）'],
    ['化龙-汇聚交换机', '化龙', '汇聚交换机', '1-9, 29-33', '同上'],
    ['化龙-OA交换机', '化龙', 'OA交换机', '1-9, 29-33', '同上'],
    ['化龙-存储交换机', '化龙', '存储交换机', '1-9, 29-33', '同上'],
    ['知识城-防火墙', '知识城', '防火墙', '1-15', '同化龙防火墙（设备IP不同）'],
    ['知识城-核心交换机', '知识城', '核心交换机', '1-9, 22-28', '通用1-9 + 知识城核心专用22-28（OSPF多实例/VRRPvrid203/M-LAG/STP/VLAN/链路聚合/Track）'],
    ['知识城-接入交换机', '知识城', '接入交换机', '1-9, 29-33', '通用1-9 + 非核心通用29-33'],
    ['知识城-IDC交换机', '知识城', 'IDC区交换机', '1-9, 29-33', '同上'],
    ['知识城-OA交换机', '知识城', 'OA交换机', '1-9, 29-33', '同上'],
    ['知识城-存储交换机', '知识城', '存储交换机', '1-9, 29-33', '同上'],
    ['知识城-上行交换机', '知识城', '上行交换机', '1-9, 29-33', '同上'],
]

for i, row_data in enumerate(groups3):
    write_row(ws3, i+2, row_data)

set_col_widths(ws3, [20, 8, 14, 18, 60])
ws3.freeze_panes = 'A2'

# ============================================================
# 保存
# ============================================================
output_path = '巡检项阈值配置表_待审核.xlsx'
wb.save(output_path)
print(f'OK: {output_path} generated')
print('  Sheet1: 巡检项阈值配置表（33项，含 parser_config/checker_config 分列）')
print('  Sheet2: 需新增Checker清单（14个）')
print('  Sheet3: 设备分组与CheckItem绑定关系（13组）')
print()
print('请打开 Excel，在黄色列"现网阈值（待填）"中填写实际指标。')