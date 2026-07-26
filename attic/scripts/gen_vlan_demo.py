# -*- coding: utf-8 -*-
"""生成VLAN检查对比演示Excel"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = Workbook()
ws = wb.active
ws.title = 'VLAN基线对比演示'

H = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
HF = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
W = Alignment(wrap_text=True, vertical='top')
T = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
DF = Font(name='微软雅黑', size=10)

hdrs = ['场景', '命令', 'parser', 'checker', 'checker_config',
        '基线输出（last）', '当前输出（current）', '预期结果', '说明']
for c,h in enumerate(hdrs,1):
    cl=ws.cell(row=1,column=c,value=h);cl.fill=H;cl.font=HF;cl.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True);cl.border=T

widths = [15, 30, 8, 10, 25, 40, 40, 12, 30]
from openpyxl.utils import get_column_letter
for i,w in enumerate(widths,1): ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = 'A2'

# ---- 场景1: 接入交换机 VLAN 不变 ----
vlan_normal = """Brief information about all VLANs:
Supported Minimum VLAN ID: 1
Supported Maximum VLAN ID: 4094
Default VLAN ID: 1
VLAN ID   Name                             Port
1         VLAN 0001                        GE1/0/41...
100       VLAN 0100                        BAGG100...
101       VLAN 0101                        BAGG100..."""
vlan_changed = """Brief information about all VLANs:
Supported Minimum VLAN ID: 1
Supported Maximum VLAN ID: 4094
Default VLAN ID: 1
VLAN ID   Name                             Port
1         VLAN 0001                        GE1/0/41...
100       VLAN 0100                        BAGG100...
101       VLAN 0101                        BAGG100...
200       VLAN 0200                        BAGG100..."""

items = [
    ('接入VLAN不变', 'display vlan brief', 'raw', 'baseline',
     '{"similarity":1.0}', vlan_normal, vlan_normal, '正常', '基线一致，无差异'),
    ('接入VLAN新增', 'display vlan brief', 'raw', 'baseline',
     '{"similarity":1.0}', vlan_normal, vlan_changed, '异常', '新增VLAN200→difflib标记'),
    ('核心VLAN删除', 'display vlan brief', 'raw', 'baseline',
     '{"similarity":1.0}',
     'VLAN 100/101/200 all present', 'VLAN 100/101 only',
     '异常', '缺少VLAN200→difflib标记'),
]

for i,r in enumerate(items):
    row = i+2
    for c,v in enumerate(r,1):
        cl=ws.cell(row=row,column=c,value=v);cl.font=DF;cl.alignment=W;cl.border=T

wb.save('VLAN检查对比演示.xlsx')
print('OK: VLAN检查对比演示.xlsx')