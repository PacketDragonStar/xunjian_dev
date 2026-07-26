# -*- coding: utf-8 -*-
"""生成所有检查项的check逻辑详解Excel"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = 'Check逻辑详解'

H = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
HF = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
W = Alignment(wrap_text=True, vertical='top')
T = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
DF = Font(name='微软雅黑', size=10)
G = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
Y = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')

hdrs = ['序号','检查项','命令','parser','parser做什么','checker','checker怎么做','判断逻辑详解','正常条件','异常条件','示例']
for c,h in enumerate(hdrs,1):
    cl=ws.cell(row=1,column=c,value=h);cl.fill=H;cl.font=HF;cl.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True);cl.border=T

widths = [5,16,28,8,28,10,30,50,25,30,30]
for i,w in enumerate(widths,1): ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = 'A2'

items = [
    # 硬件状态
    (1,'风扇状态','display fan','raw','不解parse，直接传原始文本给checker','custom',
     '调用check_fan()自定义函数','统计Normal出现次数≥1 且 无Abnormal/Fault → 正常。采集为空时报异常避免漏检。',
     'Normal≥1，无Abnormal/Fault','Normal=0或无异常关键字 → 异常','Fan1Normal,Fan2Normal→✅'),
    (2,'电源状态','display power','raw','同上','custom',
     '调用check_power()自定义函数','统计Normal出现次数≥1 且 无Abnormal/Fault/Failed/Off → 正常。',
     'Normal≥1，无异常关键字','Normal=0或异常关键字 → 异常','Power1Normal,Power2Absent→✅'),
    (3,'温度环境','display environment','raw','同上','custom',
     '调用check_env()自定义函数','从输出中提取温度数值，与temp_warning阈值(默认60°C)比较。超限或存在Fault/Abnormal→异常。',
     '各传感器温度<阈值且无Fault','温度>阈值 或 有Fault关键字','热点22°C,阈值60→✅'),
    (4,'单板状态','display device','raw','同上','custom',
     '调用check_device()自定义函数','统计Normal出现次数≥1 且 无Fault/Abnormal → 正常。',
     'Normal≥1，无Fault','Normal=0或Fault/Abnormal → 异常','Slot1MasterNormal→✅'),
    (5,'光模块收发光','display transceiver diagnosis interface','raw','同上','custom',
     '调用check_transceiver()自定义函数','逐接口提取Temp/Voltage/Bias/RX/TX当前值，与设备自带的Alarm High/Low阈值比对。越限即异常。',
     '所有参数的当前值在Alarm High/Low范围内','任一参数的当前值<Low或>High → 异常','T1/0/49RX=-3.51,Low=-11.9→✅'),
    # 接口&链路
    (6,'接口状态','display interface brief','raw','同上','baseline',
     '用difflib.SequenceMatcher逐行比对基线vs当前','第二次巡检和基线做文本diff。新增DOWN端口、UP变DOWN均高亮显示。',
     '文本完全一致(similarity=1.0)','任何变化 → 基线对比页面高亮差异','GE1/0/5从UP变DOWN→difflib标记'),
    (7,'链路聚合','display link-aggregation summary','raw','同上','custom',
     '调用check_agg()自定义函数','统计Unselected关键字出现次数。>0即异常。',
     'Unselected=0','Unselected>0 → 异常','BAGG100Selected=4,Unselected=0→✅'),
    # 二层状态
    (8,'STP状态','display stp brief','raw','同上','custom',
     '调用check_stp()自定义函数','根据root_expected参数："非根桥"→检查ROOT端口；"本端"→检查本地有ROOT标记。阻塞端口、TC/TCN拓扑变更均检测。',
     '角色正确+FWD+无阻塞+无TC','根桥角色不对/端口阻塞/TC出现→异常','BAGG100ROOT+DESI全FWD→✅'),
    (9,'VLAN清单','display vlan brief','raw','同上','baseline',
     '用difflib.SequenceMatcher比对基线','两次巡检输出完全一致→正常。VLAN增删均可发现。',
     '文本完全一致','新增或删除任何VLAN→标记差异行','新增VLAN200→difflib标记+200'),
    # 路由&协议
    (10,'OSPF邻居','display ospf peer','raw','同上','custom',
     '调用check_ospf_peer()自定义函数','统计Full关键字出现次数，与expected_full_count对比。支持多实例模式(instances参数)。',
     'Full数==期望值','Full数!=期望 或 Full=0→异常','Full=2,期望2→✅'),
    (11,'BGP邻居','display bgp peer ipv4','raw','同上','custom',
     '调用check_bgp_peer()自定义函数','优先提取"Peers in established state: N"汇总行；若无汇总行则统计Established关键字次数。与expected_established对比。',
     'Established数==期望值','Established数!=期望 或 无Established→异常','Peers in established state:1,期望1→✅'),
    (12,'VRRP状态','display vrrp brief','raw','同上','custom',
     '调用check_vrrp()自定义函数','统计Master关键字次数与vrrp_master期望值对比。检测Initialize状态。',
     'Master数==期望，无Initialize','Master数!=期望 或 有Initialize→异常','Master=1,期望1→✅'),
    (13,'路由表','display ip routing-table all-vpn-instance','raw','同上','baseline',
     '用difflib.SequenceMatcher全量比对','全VPN实例路由表与基线逐行diff。路由新增/删除/下一跳变化都可发现。',
     '文本完全一致','新增/删除路由 或 下一跳变化→标记差异','默认路由Nexthop从126变127→标记diff'),
    # 防火墙
    (14,'安全域成员','display security-zone','raw','同上','baseline',
     '用difflib.SequenceMatcher比对基线','两次输出完全一致→正常。zone增删、接口变化都可发现。',
     '文本完全一致','zone增删/接口变化→标记差异行','Trust新增一个接口→difflib标记+'),
    (15,'安全策略命中','display security-policy statistics','raw','同上','baseline',
     '用difflib.SequenceMatcher比对基线','各rule命中计数与基线diff。计数归零或暴增均可见。',
     '文本完全一致','命中计数归零/暴增→标记差异','某rule从10000→0→标记差异'),
    (16,'安全策略规则','display security-policy ip rule all','raw','同上','baseline',
     '用difflib.SequenceMatcher比对基线','策略规则全文与基线diff。规则增删改均可见。',
     '文本完全一致','策略增删/改→标记差异','新增一条policy规则→标记+'),
    (17,'RBM双机热备','display remote-backup-group status','raw','同上','custom',
     '调用check_rbm()自定义函数','从Local信息块提取Device running status(Active/Standby)、Data channel state(Up)、Control channel(Connected)。对端同理。',
     '本端Active/Standby+通道Up/Connected',
     '状态异常/通道Down/不通→异常','本端Active,数据Up,控制Connected→✅'),
    (18,'会话表','display session table ipv4','raw','同上','custom',
     '调用check_session()自定义函数','提取所有数字，与max_sessions阈值(默认500000)比较。超过则异常。',
     '所有数字<max_sessions','任意数字>max_sessions→异常','Total 4683,上限500000→✅'),
    # 高可用
    (19,'M-LAG状态','display m-lag summary','raw','同上','custom',
     '调用check_mlag()自定义函数','搜索M-LAG error/MAD conflict/M-LAG fault关键字。存在即异常。',
     '无异常关键字','有error/conflict/fault→异常','无error关键字→✅'),
    (20,'NQA探测','display nqa result','raw','同上','custom',
     '调用check_nqa()自定义函数','搜索failed/Timeout/Unreachable关键字。存在即异常。',
     '无failed/Timeout','有failed/Timeout/不可达→异常','无failed关键字→✅'),
    (21,'Track状态','display track','raw','同上','custom',
     '调用check_track()自定义函数','统计Positive次数与expected_tracks对比；检测Negative关键字。',
     '全Positive，数量==期望','有Negative 或 数量<期望→异常','Positive=12,期望12→✅'),
    (22,'IRF堆叠状态','display irf','raw','同上','custom',
     '调用check_irf()自定义函数','检查IRF mode=normal。统计唯一MemberID数量和Master数量。',
     'mode=normal,Master=1,无脑裂','脑裂(Master>1)/无Master→异常','Member1Master,Member2Standby→✅'),
    # 性能
    (23,'CPU利用率','display cpu-usage','regex','正则提取数字，cast=float','threshold',
     'check_threshold()内置函数','正则pattern="(\\d+)%"提取CPU百分比，与warning阈值(80%)+operator(<)比较。',
     '提取值<80%→正常','提取值≥80%→异常','75%→✅'),
    (24,'内存利用率','display memory','regex','正则提取FreeRatio，cast=float','threshold',
     'check_threshold()内置函数','正则pattern提取Mem:行百分比列(FreeRatio)，与warning(20%)+operator(>)比较。',
     'FreeRatio>20%→正常','FreeRatio≤20%→异常','FreeRatio=55.9%→✅'),
    # 日志
    (25,'日志缓冲','display logbuffer','strip_ts','去Comware时间戳(%Jun24...2026)','baseline',
     '去时间戳后用difflib比对基线','日志输出含大量变化的时间戳。先去时间戳再difflib比对(降低误报)。',
     '去时间戳后文本完全一致','去时间戳后仍有差异(真正的日志变化)→标记','Overwritten messages从41462→49982→标记差异'),
    # 采集/对比项
    (26,'NTP状态','display ntp status','raw','同上','contains',
     'check_contains()内置函数，检查输出是否包含指定关键字','搜索"synchronized"关键字。存在→正常。',
     '含"synchronized"','不含"synchronized"→异常','Clock status:synchronized→✅'),
    (27,'系统稳定状态','display system stable state','raw','同上','custom',
     '调用check_system_stable()自定义函数','检查System state=Stable和Redundancy state=Stable。逐Slot检查State无Fault/Abnormal。',
     'System/Redundancy全Stable','存在Stable以外的状态→异常','System state:Stable→✅'),
    # 基线对比项（全为baseline）
    (28,'版本信息','display version','raw','同上','baseline',
     '用difflib.SequenceMatcher比对基线','每次巡检保存版本信息，与基线对比。版本变更(升级/降级)可发现。',
     '完全一致','版本号/补丁/启动时间变化→标记差异','版本从1125→1200→标记差异'),
    (29,'配置备份','display current-configuration','raw','同上','baseline',
     '用difflib.SequenceMatcher全量比对','全量配置与基线diff。任意配置变更均可发现。',
     '完全一致','配置增删/改→标记差异行','新增vlan配置行→标记+'),
    (30,'LLDP邻居','display lldp neighbor-information list','raw','同上','baseline',
     '用difflib.SequenceMatcher比对基线','邻居表变化(新增/丢失/变更)可发现。',
     '完全一致','邻居增删/变化→标记差异','邻居MAC变化→标记差异'),
    (31,'错包入向','display counters inbound interface','raw','同上','baseline',
     '用difflib.SequenceMatcher比对基线','两次巡检入向接口错误计数器diff。增量不为0则异常。',
     '完全一致','错包计数增加→标记差异行','入向错包增100→标记+100'),
    (32,'错包出向','display counters outbound interface','raw','同上','baseline',
     '用difflib.SequenceMatcher比对基线','同上，出向接口错误计数器diff。',
     '完全一致','错包计数增加→标记差异行','出向错包增50→标记+50'),
    (33,'Flash空间','dir flash:/','raw','同上','baseline',
     '用difflib.SequenceMatcher比对基线','Flash空间占用变化发现。',
     '完全一致','占用率>上次→标记差异行','新增core文件→标记+'),
    (34,'ARP冲突','display arp user-ip-conflict record','raw','同上','custom',
     '调用check_arp()自定义函数','检查conflict关键字。存在"0 conflict"或"no...conflict"→正常。',
     '无冲突记录或conflict=0','有conflict记录→异常','Total:0 conflict records→✅'),
]

row = 2
for i,r in enumerate(items):
    for c,v in enumerate(r,1):
        cl=ws.cell(row=row,column=c,value=v);cl.font=DF;cl.alignment=W;cl.border=T
        if c in (5,7,8): cl.fill = G
    row+=1

wb.save('Check逻辑详解.xlsx')
print('OK: Check逻辑详解.xlsx (34条检查项)')