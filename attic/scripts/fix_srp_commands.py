# -*- coding: utf-8 -*-
"""srp 路由器最终版巡检命令"""
import os
import openpyxl
from openpyxl.utils import get_column_letter

HL_EXCEL = r'C:\Users\ZSS\Desktop\化龙\化龙\化龙配置\network_inspection\化龙设备清单_带巡检命令.xlsx'
OUT_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT = os.path.join(OUT_DIR, '化龙设备清单_已填充命令_v2.xlsx')

SRP_COMMANDS = (
    'display cpu-usage\n'
    'display memory\n'
    'display fan\n'
    'display power\n'
    'display environment\n'
    'display device\n'
    'display interface brief\n'
    'display logbuffer\n'
    'display ospf peer\n'
    'display bgp peer\n'
    'display vrrp\n'
    'display ip routing-table all-vpn-instance\n'
    'display version\n'
    'display current-configuration\n'
    'display ntp status\n'
    'display counters inbound interface\n'
    'display counters outbound interface\n'
    'display lldp neighbor-information list\n'
    'display transceiver diagnosis interface\n'
    'display system stable state\n'
    'dir flash:/'
)

print(f'srp 路由器命令 ({SRP_COMMANDS.count(chr(10)) + 1} 条):')
for i, c in enumerate(SRP_COMMANDS.split('\n'), 1):
    print(f'  {i:2d}. {c}')

# 基于之前已填充的 v1 文件修改 srp 行（如果存在）
v1_path = os.path.join(OUT_DIR, '化龙设备清单_已填充命令.xlsx')
if os.path.exists(v1_path):
    source = v1_path
else:
    source = HL_EXCEL

wb = openpyxl.load_workbook(source)
ws = wb.active

headers = [str(c.value or '').strip() for c in ws[1]]
cmd_col = name_col = None
for i, h in enumerate(headers):
    if '命令' in h: cmd_col = i
    if '设备名称' in h or '设备名' in h: name_col = i

for row in range(2, ws.max_row + 1):
    name = str(ws.cell(row=row, column=name_col + 1).value or '').strip()
    if name.lower().startswith('srp'):
        ws.cell(row=row, column=cmd_col + 1).value = SRP_COMMANDS
        print(f'\n✅ {name} → 路由器 (21 条)')

ws.column_dimensions[get_column_letter(cmd_col + 1)].width = 60
wb.save(OUTPUT)
wb.close()
print(f'\n📁 已保存: {OUTPUT}')