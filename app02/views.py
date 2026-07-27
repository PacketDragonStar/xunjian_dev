from django.http import JsonResponse
from app02.models import (
    NewDevice, DeviceGroup, CheckItem, CheckSet,
    XunjianRecord, CheckResult, AnomalyRecord, XunjianTask
)
from django.views.decorators.csrf import csrf_exempt
import pandas as pd
import difflib
import json
from django.utils.safestring import mark_safe
from django.contrib import messages
from django.shortcuts import render, redirect
import datetime
import threading
from django.db import close_old_connections
from django.conf import settings
import os
from app02.utils.pagination import Pagination
import logging
import pkgutil
import importlib
from concurrent.futures import ThreadPoolExecutor, as_completed

base_dir = settings.BASE_DIR
file_path = os.path.join(base_dir, 'app02', 'static', 'example.xlsx')


# ── 登录 ─────────────────────────────────────────────────
def login_view(request):
    """用户登录"""
    from app02.forms import LoginForm
    if request.method == 'GET':
        form = LoginForm()
        return render(request, 'login.html', {'form': form})
    form = LoginForm(data=request.POST)
    if form.is_valid():
        from django.contrib import auth
        user = auth.authenticate(
            request,
            username=form.cleaned_data['username'],
            password=form.cleaned_data['password'],
        )
        if user:
            auth.login(request, user)
            request.session['info'] = {'name': user.username, 'id': user.id}
            return redirect('/')
        form.add_error(None, '用户名或密码错误')
    return render(request, 'login.html', {'form': form})




from app02.engine.executor import run_xunjian as _run_xunjian_new
from app02.engine.executor import _build_conn_kwargs
from app02.engine.capability import ensure_capabilities, PROBE_COMMAND
from netmiko import ConnectHandler

_new_logger = logging.getLogger('app02.new_engine')


# ── 新版：触发巡检 ─────────────────────────────────────────
@csrf_exempt
def new_run_xunjian(request):
    """新版巡检执行接口（后台线程异步执行，立即返回 task_id）"""
    if request.method != 'POST':
        return JsonResponse({'status': False, 'error': '仅支持POST'}, status=405)

    operator   = request.session.get('info', {}).get('name', '未知')
    device_ids = request.POST.getlist('device_ids')
    device_ids = [int(i) for i in device_ids if i] or None
    checkset_id = request.POST.get('checkset_id')
    checkset_id = int(checkset_id) if checkset_id else None

    # 计算将巡检的设备总数（用于进度条初始值；无分组的设备会被引擎跳过）
    qs = NewDevice.objects.filter(enabled=True)
    if device_ids:
        qs = qs.filter(id__in=device_ids)
    elif checkset_id:
        checkset_obj = CheckSet.objects.filter(id=checkset_id, enabled=True).prefetch_related('groups').first()
        if checkset_obj:
            group_ids = list(checkset_obj.groups.values_list('id', flat=True))
            qs = qs.filter(group__id__in=group_ids)
    device_count = qs.count()
    if device_count == 0:
        return JsonResponse({'status': False, 'error': '没有可巡检的设备'}, status=400)

    # 先建任务记录（排队中），再由后台线程执行，避免阻塞 HTTP 请求
    task = XunjianTask.objects.create(
        status='queued', operator=operator,
        checkset_id=checkset_id, device_count=device_count,
    )

    def _worker():
        try:
            res = _run_xunjian_new(
                operator=operator,
                device_ids=device_ids,
                checkset_id=checkset_id,
                task_id=task.id,
            )
            XunjianTask.objects.filter(id=task.id).update(
                xunjian_time=res['time'], result=res['result'],
            )
            # 巡检后 hook pipeline：sync_cmdb / detect_capabilities ...
            from app02.engine.post_inspection import run_post_inspection_hooks
            run_post_inspection_hooks(task.id, res['time'], operator)
        except Exception as e:
            close_old_connections()
            XunjianTask.objects.filter(id=task.id).update(
                status='failed', error=str(e)[:500],
            )
            _new_logger.error(f'新版巡检执行失败(task={task.id}): {e}')
        finally:
            close_old_connections()

    t = threading.Thread(target=_worker, daemon=True)
    t.start()
    return JsonResponse({'status': True, 'task_id': task.id})


# ── 新版：历史记录列表 ──────────────────────────────────────
def new_search_history(request):
    """新版巡检历史列表"""
    search   = request.GET.get('q', '')
    date_from = request.GET.get('date_from', '')
    date_to   = request.GET.get('date_to', '')
    qs = XunjianRecord.objects.all()
    if search:
        qs = qs.filter(time__icontains=search) | \
             qs.filter(operator__icontains=search) | \
             qs.filter(result__icontains=search)
    if date_from:
        qs = qs.filter(time__gte=date_from)
    if date_to:
        # date_to 加一天，让当天的记录也能被包含
        qs = qs.filter(time__lte=date_to + ' 23:59:59')
    qs = qs.order_by('-time')
    baseline = XunjianRecord.objects.filter(is_baseline=True).first()
    page_obj = Pagination(request, qs)
    return render(request, 'new_search_history.html', {
        'data_list':   page_obj.page_queryset,
        'page_string': page_obj.html(),
        'search_data': search,
        'date_from':   date_from,
        'date_to':     date_to,
        'baseline':    baseline,
    })


# ── 登出 ─────────────────────────────────────────────────
def logout_view(request):
    """用户登出"""
    from django.contrib import auth
    auth.logout(request)
    request.session.clear()
    return redirect('/login/')


# ── 新版：巡检详情（异常列表）──────────────────────────────
def new_history_detail(request):
    """新版巡检详情：异常列表"""
    xunjian_time = request.GET.get('time')
    tab = request.GET.get('tab', 'anomalies')  # 'anomalies' | 'raw'
    if not xunjian_time:
        return redirect('/new/history/')
    anomalies = AnomalyRecord.objects.filter(time=xunjian_time).order_by('device')
    record    = XunjianRecord.objects.filter(time=xunjian_time).first()
    if record is None:
        return redirect('/new/history/')
    # 按设备分组原始输出，供「回显查看」使用
    from collections import defaultdict
    raw_results = CheckResult.objects.filter(time=xunjian_time).order_by('device', 'command')
    device_results = defaultdict(list)
    all_cmds = set()
    for r in raw_results:
        device_results[r.device].append(r)
        all_cmds.add(r.command)
    device_results = dict(device_results)
    all_commands = sorted(all_cmds)
    anomaly_set = set((a.device, a.command) for a in anomalies)
    return render(request, 'new_info_history.html', {
        'anomalies':      anomalies,
        'record':         record,
        'time':           xunjian_time,
        'tab':            tab,
        'tag':            '' if (anomalies.exists() or record.failed_devices > 0) else '本次巡检无异常',
        'device_results': device_results,
        'all_commands':   all_commands,
        'device_names':   sorted(device_results.keys()),
        'anomaly_set':    anomaly_set,
    })


# ── 新版：删除巡检记录 ──────────────────────────────────────
def new_history_delete(request):
    xunjian_time = request.GET.get('time')
    XunjianRecord.objects.filter(time=xunjian_time).delete()
    CheckResult.objects.filter(time=xunjian_time).delete()
    AnomalyRecord.objects.filter(time=xunjian_time).delete()
    return redirect('/new/history/')


# ── 新版：设置基线 ──────────────────────────────────────────
@csrf_exempt
def new_set_baseline(request):
    xunjian_time = request.GET.get('time')
    XunjianRecord.objects.filter(is_baseline=True).update(is_baseline=False)
    XunjianRecord.objects.filter(time=xunjian_time).update(is_baseline=True)
    return redirect('/new/history/')


# ── 新版：确认单条异常 ──────────────────────────────────────
@csrf_exempt
def new_confirm_notes(request):
    xunjian_time = request.GET.get('time')
    device       = request.GET.get('device')
    command      = request.GET.get('command')
    AnomalyRecord.objects.filter(
        time=xunjian_time, device=device, command=command
    ).update(confirm=True)
    _new_update_overall(xunjian_time)
    return redirect(f'/new/history/detail/?time={xunjian_time}')


# ── 新版：确认全部异常 ──────────────────────────────────────
@csrf_exempt
def new_confirm_all(request):
    xunjian_time = request.GET.get('time')
    AnomalyRecord.objects.filter(time=xunjian_time).update(confirm=True)
    XunjianRecord.objects.filter(time=xunjian_time).update(result='正常')
    return redirect(f'/new/history/detail/?time={xunjian_time}')


def _new_update_overall(xunjian_time):
    """若无未确认异常则更新总记录为正常"""
    if not AnomalyRecord.objects.filter(time=xunjian_time, confirm=False).exists():
        XunjianRecord.objects.filter(time=xunjian_time).update(result='正常')


# ── 巡检历史快速浏览（JSON API） ─────────────────────────────
def xunjian_history_browse(request):
    """返回某次巡检的设备列表和命令列表，供前端快速浏览"""
    xunjian_time = request.GET.get('time', '')
    if not xunjian_time:
        return JsonResponse({'status': False, 'error': '缺少参数 time'})
    results = CheckResult.objects.filter(time=xunjian_time).order_by('device', 'command')
    if not results.exists():
        return JsonResponse({'status': False, 'error': '该次巡检无命令回显数据'})
    
    devices = list(results.values_list('device', flat=True).distinct())
    commands = list(results.values_list('command', flat=True).distinct())
    
    # 设备-命令映射
    device_cmd_map = {}
    for r in results:
        device_cmd_map.setdefault(r.device, []).append(r.command)
    
    return JsonResponse({
        'status': True,
        'devices': devices,
        'commands': commands,
        'device_cmd_map': device_cmd_map,
    })


@csrf_exempt
def xunjian_history_raw_output(request):
    """返回某次巡检某设备某命令的原始回显"""
    xunjian_time = request.GET.get('time', '')
    device = request.GET.get('device', '')
    command = request.GET.get('command', '')
    if not all([xunjian_time, device, command]):
        return JsonResponse({'status': False, 'error': '缺少参数'})
    
    result = CheckResult.objects.filter(
        time=xunjian_time, device=device, command=command
    ).first()
    if not result:
        return JsonResponse({'status': False, 'error': '未找到该命令回显'})
    
    # 同时取基线数据
    baseline_record = XunjianRecord.objects.filter(is_baseline=True).first()
    baseline_result = None
    if baseline_record:
        baseline_result = CheckResult.objects.filter(
            time=baseline_record.time, device=device, command=command
        ).first()
    
    return JsonResponse({
        'status': True,
        'device': device,
        'command': command,
        'current_raw': result.result,
        'baseline_raw': baseline_result.result if baseline_result else '',
        'baseline_time': baseline_record.time if baseline_record else '',
    })


# ── 配置下载 ────────────────────────────────────────────────
def config_download(request):
    """下载某次巡检某设备的配置备份为 .txt 文件"""
    from django.http import HttpResponse
    
    xunjian_time = request.GET.get('time', '')
    device = request.GET.get('device', '')
    if not xunjian_time or not device:
        return HttpResponse('缺少参数 time/device', status=400)
    
    cmd = 'display current-configuration'
    result = CheckResult.objects.filter(
        time=xunjian_time, device=device, command=cmd
    ).first()
    if not result:
        # 尝试第二个可能的命令
        result = CheckResult.objects.filter(
            time=xunjian_time, device=device, command__icontains='current-configuration'
        ).first()
    if not result:
        return HttpResponse(f'未找到 {device} 在 {xunjian_time} 的配置备份', status=404)
    
    filename = f"config_{device}_{xunjian_time.replace(':','-').replace(' ','_')}.txt"
    response = HttpResponse(result.result, content_type='text/plain; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def _strip_for_compare(text, cfg):
    """按 CheckItem.compare_strip 配置清洗文本：去前 N 行 + 跳过匹配正则的行。"""
    if not cfg or not text:
        return text
    import re as _re
    lines = text.splitlines()
    head = int(cfg.get('head_lines', 0) or 0)
    if head > 0:
        lines = lines[head:]
    pats = cfg.get('skip_patterns') or []
    if pats:
        rxs = [_re.compile(p) for p in pats]
        lines = [ln for ln in lines if not any(rx.search(ln) for rx in rxs)]
    return '\n'.join(lines)


# ── 新版：命令输出任意批次对比 ────────────────────────────────
def new_text_compare(request):
    """对任意设备/命令的命令回显做两批次对比。
    默认：左(基准)=基线(若未设基线则取最新批次)，右(目标)=最新批次。
    兼容旧入口：time 参数视为右批次、compare_time 视为左批次。
    对比前可按 CheckItem.compare_strip 清洗文本（去表头/过滤行，可开关）。
    若 CheckItem.extract_parser 已配置，额外展示结构化提取字段（仅展示）。
    """
    from app02.engine.pipeline import EXTRACTORS

    command = request.GET.get('command', '').strip()
    device = request.GET.get('device', '').strip()
    right_time = request.GET.get('right_time') or request.GET.get('time', '')
    left_time = request.GET.get('left_time') or request.GET.get('compare_time', '')

    # 所有巡检批次（用于下拉）
    records = list(XunjianRecord.objects.all().order_by('-time'))
    baseline = next((r for r in records if r.is_baseline), None)

    # 默认批次：左=基线(或最新)，右=最新
    if not left_time:
        left_time = baseline.time if baseline else (records[0].time if records else '')
    if not right_time:
        right_time = records[0].time if records else ''

    # 下拉选项来源（来自已采集数据）
    devices = list(CheckResult.objects.values_list('device', flat=True).distinct().order_by('device'))
    commands = list(CheckResult.objects.values_list('command', flat=True).distinct().order_by('command'))

    # 取两侧数据
    left = (CheckResult.objects.filter(time=left_time, device=device, command=command).first()
            if (device and command) else None)
    right = (CheckResult.objects.filter(time=right_time, device=device, command=command).first()
             if (device and command) else None)

    # 对比清洗配置
    check_item = CheckItem.objects.filter(command=command).first()
    strip_cfg = check_item.compare_strip if check_item else None
    strip_on = (request.GET.get('strip', '1') != '0') and bool(strip_cfg)

    left_raw = left.result if (left and left.result is not None) else ''
    right_raw = right.result if (right and right.result is not None) else ''
    left_diff = _strip_for_compare(left_raw, strip_cfg) if strip_on else left_raw
    right_diff = _strip_for_compare(right_raw, strip_cfg) if strip_on else right_raw

    # 字段提取（仅展示，不影响判定）
    extract_parser = check_item.extract_parser if check_item else ''
    left_extract = {}
    right_extract = {}
    if extract_parser and extract_parser in EXTRACTORS:
        left_extract = EXTRACTORS[extract_parser](left_raw)
        right_extract = EXTRACTORS[extract_parser](right_raw)

    diff_table = ''
    if left and right:
        left_lines = left_diff.splitlines()
        right_lines = right_diff.splitlines()
        total_lines = max(len(left_lines), len(right_lines))
        use_context = total_lines > 500
        diff_table = difflib.HtmlDiff(wrapcolumn=120).make_table(
            left_lines,
            right_lines,
            fromdesc='基准: ' + (left_time or '-'),
            todesc='目标: ' + (right_time or '-'),
            context=use_context,
        )
        if use_context:
            from django.utils.html import format_html
            diff_table = str(format_html(
                '<div class="alert alert-info">输出行数过多({}行)，已自动切换为上下文模式（仅显示变化行附近内容）</div>',
                total_lines
            )) + diff_table

    show_result = bool(device and command and (left or right))

    ctx = {
        'records': records,
        'baseline_time': baseline.time if baseline else '',
        'devices': devices,
        'commands': commands,
        'device': device,
        'command': command,
        'left_time': left_time,
        'right_time': right_time,
        'left_result': left_raw,
        'right_result': right_raw,
        'left_extract': list(left_extract.items()),
        'right_extract': list(right_extract.items()),
        'extract_parser': extract_parser,
        'compare_strip': strip_cfg,
        'strip_on': strip_on,
        'left_missing': bool(device and command and not (left and left.result is not None)),
        'right_missing': bool(device and command and not (right and right.result is not None)),
        'diff_table': mark_safe(diff_table) if diff_table else '',
        'show_result': show_result,
    }
    return render(request, 'new_text_compare.html', ctx)


# ── 新版：巡检项管理 ────────────────────────────────────────
def new_checkitem_list(request):
    search = request.GET.get('q', '')
    qs = CheckItem.objects.all()
    if search:
        qs = qs.filter(name__icontains=search) | qs.filter(command__icontains=search)
    page_obj = Pagination(request, qs)
    return render(request, 'new_checkitem_list.html', {
        'queryset':    page_obj.page_queryset,
        'page_string': page_obj.html(),
        'search_data': search,
    })


@csrf_exempt
def new_checkitem_add(request):
    if request.method != 'POST':
        return JsonResponse({'status': False, 'error': '仅支持POST'})
    data = request.POST
    try:
        p_conf = json.loads(data.get('parser_config', 'null') or 'null')
        c_conf = json.loads(data.get('checker_config', 'null') or 'null')
    except json.JSONDecodeError:
        p_conf, c_conf = None, None
    name    = data.get('name', '').strip()
    command = data.get('command', '').strip()
    if not name or not command:
        return JsonResponse({'status': False, 'error': '名称和命令不能为空'})
    try:
        c_strip = data.get('compare_strip', '').strip()
        c_strip_val = None
        if c_strip:
            try:
                c_strip_val = json.loads(c_strip)
            except json.JSONDecodeError:
                c_strip_val = None
        CheckItem.objects.create(
            name=name, command=command,
            parser=data.get('parser', 'raw'), parser_config=p_conf,
            checker=data.get('checker', 'baseline'), checker_config=c_conf,
            error_note=data.get('error_note', '请检查'),
            timeout=int(data.get('timeout', 30) or 30),
            enabled=data.get('enabled', 'true') == 'true',
            extract_parser=data.get('extract_parser', '') or '',
            compare_strip=c_strip_val,
            feature=data.get('feature', 'base') or 'base',
        )
    except Exception as e:
        return JsonResponse({'status': False, 'error': str(e)})
    return JsonResponse({'status': True})


@csrf_exempt
def new_checkitem_edit(request):
    uid = request.GET.get('uid')
    obj = CheckItem.objects.filter(id=uid).first()
    if not obj:
        return JsonResponse({'status': False, 'error': '不存在'})
    data = request.POST
    for f in ['name', 'command', 'parser', 'checker', 'error_note', 'extract_parser', 'feature']:
        if f in data:
            setattr(obj, f, data[f])
    if 'parser_config' in data:
        try:
            obj.parser_config = json.loads(data['parser_config'])
        except json.JSONDecodeError:
            pass
    if 'checker_config' in data:
        try:
            obj.checker_config = json.loads(data['checker_config'])
        except json.JSONDecodeError:
            pass
    if 'compare_strip' in data:
        raw = data['compare_strip'].strip()
        if raw:
            try:
                obj.compare_strip = json.loads(raw)
            except json.JSONDecodeError:
                pass
        else:
            obj.compare_strip = None
    if 'timeout' in data:
        obj.timeout = int(data['timeout'])
    if 'enabled' in data:
        obj.enabled = data['enabled'] == 'true'
    obj.save()
    return JsonResponse({'status': True})


@csrf_exempt
def new_checkitem_delete(request):
    uid = request.GET.get('uid')
    CheckItem.objects.filter(id=uid).delete()
    return JsonResponse({'status': True})


def new_checkitem_detail(request):
    uid = request.GET.get('uid')
    obj = CheckItem.objects.filter(id=uid).first()
    if not obj:
        return JsonResponse({'status': False, 'error': '不存在'})
    return JsonResponse({'status': True, 'data': {
        'id': obj.id, 'name': obj.name, 'command': obj.command,
        'parser': obj.parser, 'parser_config': obj.parser_config,
        'checker': obj.checker, 'checker_config': obj.checker_config,
        'error_note': obj.error_note, 'timeout': obj.timeout,
        'enabled': obj.enabled,
        'compare_strip': obj.compare_strip,
        'extract_parser': obj.extract_parser or '',
        'feature': obj.feature or 'base',
    }})


# ── 新版：分组管理 ──────────────────────────────────────────
def new_group_list(request):
    groups    = DeviceGroup.objects.prefetch_related('check_items').all()
    all_items = CheckItem.objects.all()
    return render(request, 'new_group_list.html', {
        'groups': groups, 'all_items': all_items
    })


@csrf_exempt
def new_group_add(request):
    if request.method != 'POST':
        return JsonResponse({'status': False, 'error': '仅支持POST'})
    name = request.POST.get('name', '').strip()
    if not name:
        return JsonResponse({'status': False, 'error': '分组名不能为空'})
    group = DeviceGroup.objects.create(
        name=name,
        description=request.POST.get('description', '')
    )
    item_ids = request.POST.getlist('check_items')
    if item_ids:
        group.check_items.set(CheckItem.objects.filter(id__in=item_ids))
    return JsonResponse({'status': True})


@csrf_exempt
def new_group_edit(request):
    uid = request.GET.get('uid')
    obj = DeviceGroup.objects.filter(id=uid).first()
    if not obj:
        return JsonResponse({'status': False, 'error': '不存在'})
    obj.name        = request.POST.get('name', obj.name)
    obj.description = request.POST.get('description', obj.description)
    obj.save()
    item_ids_raw = request.POST.getlist('check_items')
    # 过滤掉空字符串占位（前端未选时传 ''）
    item_ids = [i for i in item_ids_raw if i]
    # 只有前端明确传递了 check_items 字段时才更新
    if 'check_items' in request.POST:
        obj.check_items.set(CheckItem.objects.filter(id__in=item_ids))
    return JsonResponse({'status': True})


@csrf_exempt
def new_group_delete(request):
    uid = request.GET.get('uid')
    DeviceGroup.objects.filter(id=uid).delete()
    return JsonResponse({'status': True})


# ── 新版：新设备管理 ────────────────────────────────────────
def new_device_list(request):
    search = request.GET.get('q', '')
    qs = NewDevice.objects.select_related('group').all()
    if search:
        qs = qs.filter(name__icontains=search) | qs.filter(ip__icontains=search)
    page_obj = Pagination(request, qs)
    # 设备分类 2.0：为每个设备预计算能力感知展示信息，供模板渲染
    from app02.models import DEVICE_CLASS_CHOICES
    dc_label = dict(DEVICE_CLASS_CHOICES)
    devs = list(page_obj.page_queryset)             # 物化，避免模板二次迭代重新查库丢失属性
    for d in devs:
        extra = d.extra or {}
        d.dc_label = dc_label.get(d.device_class, d.device_class)
        caps = extra.get('capabilities')             # list 或 None
        d.caps_none = caps is None                  # 从未检测
        d.caps_list = caps or []                    # [] 表示已检测确无特性
        d.caps_ts = extra.get('capabilities_ts')
        d.protocol = bool(extra.get('protocol_inspection'))
        d.pending_list = extra.get('pending_capabilities') or []
    return render(request, 'new_device_list.html', {
        'queryset':    devs,
        'page_string': page_obj.html(),
        'groups':      DeviceGroup.objects.all(),
        'search_data': search,
    })


@csrf_exempt
def new_device_add(request):
    if request.method != 'POST':
        return JsonResponse({'status': False, 'error': '仅支持POST'})
    data = request.POST
    name        = data.get('name', '').strip()
    ip          = data.get('ip', '').strip()
    device_type = data.get('device_type', '').strip()
    username    = data.get('username', '').strip()
    password    = data.get('password', '').strip()
    if not name or not ip or not device_type or not username:
        return JsonResponse({'status': False, 'error': '设备名、IP、设备类型、用户名不能为空'})
    try:
        extra = json.loads(data.get('extra', '{}') or '{}')
    except json.JSONDecodeError:
        extra = {}
    group_id = data.get('group', '').strip()
    group = DeviceGroup.objects.filter(id=group_id).first() if group_id else None
    try:
        NewDevice.objects.create(
            name=name, ip=ip, group=group,
            device_type=device_type,
            username=username, password=password,
            extra=extra
        )
    except Exception as e:
        return JsonResponse({'status': False, 'error': str(e)})
    return JsonResponse({'status': True})


@csrf_exempt
def new_device_edit(request):
    uid = request.GET.get('uid')
    obj = NewDevice.objects.filter(id=uid).first()
    if not obj:
        return JsonResponse({'status': False, 'error': '设备不存在'})
    data = request.POST
    try:
        extra = json.loads(data.get('extra', '{}') or '{}')
    except json.JSONDecodeError:
        extra = obj.extra
    group_id = data.get('group', '').strip()
    group = DeviceGroup.objects.filter(id=group_id).first() if group_id else obj.group
    obj.name        = data.get('name', obj.name)
    obj.ip          = data.get('ip', obj.ip)
    obj.group       = group
    obj.device_type = data.get('device_type', obj.device_type)
    obj.username    = data.get('username', obj.username)
    # 密码为空时保留原密码，不覆盖
    new_pwd = data.get('password', '').strip()
    if new_pwd:
        obj.password = new_pwd
    obj.extra           = extra
    # 连接层扩展字段
    if 'conn_type' in data:
        obj.conn_type = data['conn_type']
    if 'port' in data:
        try:
            obj.port = int(data['port']) if data['port'].strip() else None
        except (ValueError, TypeError):
            obj.port = None
    if 'enable_password' in data:
        obj.enable_password = data['enable_password']
    if 'ssh_key_file' in data:
        obj.ssh_key_file = data['ssh_key_file']
    # 角色/站点
    if 'role' in data:
        obj.role = data['role']
    if 'site' in data:
        obj.site = data['site']
    obj.save()
    return JsonResponse({'status': True})


@csrf_exempt
def new_device_delete(request):
    uid = request.GET.get('uid')
    NewDevice.objects.filter(id=uid).delete()
    return redirect('/new/device/list/')


def new_device_detail(request):
    uid = request.GET.get('uid')
    obj = NewDevice.objects.filter(id=uid).first()
    if not obj:
        return JsonResponse({'status': False, 'error': '不存在'})
    return JsonResponse({'status': True, 'data': {
        'id': obj.id, 'name': obj.name, 'ip': obj.ip,
        'group': obj.group_id, 'device_type': obj.device_type,
        'username': obj.username, 'password': obj.password,
        'extra': obj.extra, 'enabled': obj.enabled,
    }})


@csrf_exempt
def new_device_capability(request):
    """设备能力感知操作的 Web 接口（与「检测能力」/「协议巡检」开关对应）。

    action:
      discover  连接设备跑探针 → 写 extra['capabilities']（纯透明，不执行）
      toggle    翻转 extra['protocol_inspection'] 开关
      set       显式设置 protocol_inspection（on=1/true 开启，其他关闭）
    """
    if request.method != 'POST':
        return JsonResponse({'status': False, 'error': '仅支持POST'})
    action = request.POST.get('action') or request.GET.get('action')

    # batch_confirm 不需要 uid，提前返回
    if action == 'batch_confirm':
        """批量确认：合并 pending → capabilities，并为所有已确认能力绑定巡检项。"""
        qs = NewDevice.objects.filter(enabled=True)
        done = 0
        items_added = 0
        for dev in qs:
            extra = dict(dev.extra or {})
            pending = extra.pop('pending_capabilities', None) or []
            existing = set(extra.get('capabilities') or [])
            if pending:
                extra['capabilities'] = list(existing | set(pending))
                dev.extra = extra
                dev.save(update_fields=['extra'])
            # 为该设备所有已确认能力绑定对应巡检项（缺失的才补）
            all_caps = list(existing | set(pending))
            if all_caps and dev.group:
                to_add = CheckItem.objects.filter(
                    enabled=True, feature__in=all_caps
                ).exclude(id__in=dev.group.check_items.values_list('id', flat=True))
                if to_add.exists():
                    dev.group.check_items.add(*to_add)
                    items_added += to_add.count()
            if pending or (all_caps and dev.group):
                done += 1
        return JsonResponse({'status': True, 'devices': done, 'items_added': items_added})

    uid = request.POST.get('uid') or request.GET.get('uid')
    obj = NewDevice.objects.filter(id=uid).first()
    if not obj:
        return JsonResponse({'status': False, 'error': '设备不存在'})

    if action == 'discover':
        conn = None
        try:
            conn = ConnectHandler(**_build_conn_kwargs(obj))
            try:
                conn.send_command('screen-length disable',
                                  expect_string=r'>|\$|#|\]', read_timeout=10)
            except Exception:
                pass
            caps = ensure_capabilities(obj, conn, force=True)
        except Exception as e:
            return JsonResponse({'status': False, 'error': f'连接/探测失败: {e}'})
        finally:
            if conn is not None:
                try:
                    conn.disconnect()
                except Exception:
                    pass
        if caps is None:
            return JsonResponse({'status': False, 'error': '探测失败（保守按基础项）'})
        # 只把新发现的写入 pending，已确认的不动
        extra = dict(obj.extra or {})
        existing = set(extra.get('capabilities') or [])
        old_pending = set(extra.get('pending_capabilities') or [])
        new = set(caps) - existing - old_pending
        if new:
            extra['pending_capabilities'] = list(old_pending | new)
            obj.extra = extra
            obj.save(update_fields=['extra'])
        return JsonResponse({'status': True,
                             'caps': caps,
                             'confirmed': list(existing),
                             'pending': extra.get('pending_capabilities') or [],
                             'new': list(new),
                             'protocol': bool(extra.get('protocol_inspection'))})

    if action == 'pending':
        """查询待确认的能力列表。"""
        extra = obj.extra or {}
        pending = extra.get('pending_capabilities') or []
        return JsonResponse({'status': True, 'pending': pending})

    if action == 'confirm':
        """确认选中的 pending 能力：移到 capabilities，并自动绑定对应巡检项到设备分组。"""
        extra = dict(obj.extra or {})
        pending = extra.pop('pending_capabilities', None) or []
        existing = set(extra.get('capabilities') or [])
        extra['capabilities'] = list(existing | set(pending))
        obj.extra = extra
        obj.save(update_fields=['extra'])
        # 自动将对应 feature 的巡检项绑定到设备所在分组
        added = 0
        if pending and obj.group:
            items_to_add = CheckItem.objects.filter(
                enabled=True, feature__in=pending
            ).exclude(id__in=obj.group.check_items.values_list('id', flat=True))
            if items_to_add.exists():
                obj.group.check_items.add(*items_to_add)
                added = items_to_add.count()
        return JsonResponse({'status': True, 'caps': extra['capabilities'],
                             'check_items_added': added})

    if action == 'dismiss':
        """关闭能力提示。"""
        extra = dict(obj.extra or {})
        extra.pop('pending_capabilities', None)
        extra['capabilities_nag_disabled'] = True
        obj.extra = extra
        obj.save(update_fields=['extra'])
        return JsonResponse({'status': True})

    if action == 'link_items':
        """从设备页直接给所在分组绑定巡检项。feature 参数可选过滤。"""
        feature = request.POST.get('feature', '').strip() or None
        item_ids = request.POST.getlist('item_ids')
        if not obj.group:
            return JsonResponse({'status': False, 'error': '该设备未绑定分组'})
        if item_ids:
            qs = CheckItem.objects.filter(id__in=item_ids, enabled=True)
            obj.group.check_items.add(*qs)
            return JsonResponse({'status': True, 'added': qs.count()})
        # 无 item_ids 时，列出可添加的巡检项（排除已绑定）
        qs = CheckItem.objects.filter(enabled=True).exclude(
            id__in=obj.group.check_items.values_list('id', flat=True)
        )
        if feature:
            qs = qs.filter(feature=feature)
        data = [{'id': it.id, 'name': it.name, 'command': it.command,
                 'feature': it.feature}
                for it in qs.order_by('feature', 'command')[:200]]
        return JsonResponse({'status': True, 'items': data})

    if action == 'bound_items':
        """查看设备所在分组已绑定的巡检项列表。"""
        if not obj.group:
            return JsonResponse({'status': False, 'error': '该设备未绑定分组'})
        items = obj.group.check_items.filter(enabled=True).order_by('feature', 'command')
        data = [{'id': it.id, 'name': it.name, 'command': it.command,
                 'feature': it.feature, 'checker': it.checker}
                for it in items]
        return JsonResponse({'status': True, 'items': data, 'group': obj.group.name})

    if action == 'unlink_items':
        """从设备所在分组移除指定巡检项。"""
        item_ids = request.POST.getlist('item_ids')
        if not obj.group:
            return JsonResponse({'status': False, 'error': '该设备未绑定分组'})
        if item_ids:
            qs = CheckItem.objects.filter(id__in=item_ids)
            obj.group.check_items.remove(*qs)
            return JsonResponse({'status': True, 'removed': qs.count()})
        return JsonResponse({'status': False, 'error': '缺少 item_ids'})

    if action in ('toggle', 'set'):
        on = True if action == 'toggle' \
            else str(request.POST.get('on', '1')).lower() in ('1', 'true', 'on')
        if action == 'toggle':
            extra = dict(obj.extra or {})
            on = not bool(extra.get('protocol_inspection'))
        else:
            extra = dict(obj.extra or {})
        extra['protocol_inspection'] = on
        obj.extra = extra
        obj.save(update_fields=['extra'])
        return JsonResponse({'status': True, 'protocol': on,
                             'caps': extra.get('capabilities')})

    return JsonResponse({'status': False, 'error': '未知 action'})



# ── 新版：检查集管理 ────────────────────────────────────────
def new_checkset_list(request):
    search = request.GET.get('q', '')
    qs = CheckSet.objects.prefetch_related('groups').all()
    if search:
        qs = qs.filter(name__icontains=search)
    return render(request, 'new_checkset_list.html', {
        'checksets':   qs,
        'all_groups':  DeviceGroup.objects.all(),
        'search_data': search,
    })


@csrf_exempt
def new_checkset_add(request):
    if request.method != 'POST':
        return JsonResponse({'status': False, 'error': '仅支持POST'})
    name = request.POST.get('name', '').strip()
    if not name:
        return JsonResponse({'status': False, 'error': '名称不能为空'})
    cs = CheckSet.objects.create(
        name=name,
        description=request.POST.get('description', ''),
    )
    group_ids = request.POST.getlist('groups')
    if group_ids:
        cs.groups.set(DeviceGroup.objects.filter(id__in=group_ids))
    return JsonResponse({'status': True})


@csrf_exempt
def new_checkset_edit(request):
    uid = request.GET.get('uid')
    obj = CheckSet.objects.filter(id=uid).first()
    if not obj:
        return JsonResponse({'status': False, 'error': '不存在'})
    obj.name        = request.POST.get('name', obj.name)
    obj.description = request.POST.get('description', obj.description)
    obj.enabled     = request.POST.get('enabled', 'true') == 'true'
    obj.save()
    group_ids_raw = request.POST.getlist('groups')
    group_ids = [i for i in group_ids_raw if i]
    if 'groups' in request.POST:
        obj.groups.set(DeviceGroup.objects.filter(id__in=group_ids))
    return JsonResponse({'status': True})


@csrf_exempt
def new_checkset_delete(request):
    uid = request.GET.get('uid')
    CheckSet.objects.filter(id=uid).delete()
    return JsonResponse({'status': True})


def new_checkset_detail(request):
    uid = request.GET.get('uid')
    obj = CheckSet.objects.filter(id=uid).prefetch_related('groups').first()
    if not obj:
        return JsonResponse({'status': False, 'error': '不存在'})
    return JsonResponse({'status': True, 'data': {
        'id':          obj.id,
        'name':        obj.name,
        'description': obj.description,
        'enabled':     obj.enabled,
        'group_ids':   list(obj.groups.values_list('id', flat=True)),
    }})


# ── 新版：巡检首页 ──────────────────────────────────────────
def new_xunjian_page(request):
    """新版巡检首页：显示分组、设备、检查集"""
    groups    = DeviceGroup.objects.prefetch_related('check_items').all()
    devices   = NewDevice.objects.select_related('group').filter(enabled=True)
    checksets = CheckSet.objects.filter(enabled=True)
    baseline = XunjianRecord.objects.filter(is_baseline=True).first()
    return render(request, 'new_xunjian_page.html', {
        'groups':    groups,
        'devices':   devices,
        'checksets': checksets,
        'baseline':  baseline,
    })


# ══════════════════════════════════════════════════════════════
# AI辅助批量导入巡检项
# ══════════════════════════════════════════════════════════════

import re as _re


def _ai_analyze_output(command: str, output: str) -> dict:
    """
    内置规则引擎：根据命令名称和输出内容，推断最合适的
    parser / checker / checker_config 组合。
    返回一个建议配置字典。
    """
    text = output.strip()
    lines = text.splitlines()
    suggestion = {
        'parser':         'raw',
        'parser_config':  None,
        'checker':        'baseline',
        'checker_config': {'similarity': 1.0},
        'reason':         '默认：原始文本基线对比',
    }

    # ── 规则1：输出只有一行且全为数字 → threshold
    if len(lines) == 1 and _re.match(r'^[\d\.]+$', text):
        suggestion.update({
            'parser':         'regex',
            'parser_config':  {'pattern': r'([\d\.]+)', 'group': 1, 'cast': 'float'},
            'checker':        'threshold',
            'checker_config': {'warning': float(text) * 1.2, 'operator': '<='},
            'reason':         f'检测到纯数值输出（{text}），建议阈值检查，警戒值已设为当前值×1.2',
        })
        return suggestion

    # ── 规则2：含时间戳行 → strip_ts
    ts_pattern = _re.compile(r'\d{4}[-/]\d{2}[-/]\d{2}[T ]\d{2}:\d{2}:\d{2}')
    ts_lines = [l for l in lines if ts_pattern.search(l)]
    if len(ts_lines) > len(lines) * 0.3:  # 超过30%行有时间戳
        suggestion.update({
            'parser':        'strip_ts',
            'parser_config': {'patterns': [r'\d{4}-\d{2}-\d{2}[T ]\d{2}:\d{2}:\d{2}(?:\.\d+)?']},
            'checker':       'baseline',
            'checker_config': {'similarity': 0.95},
            'reason':        '检测到大量时间戳，建议去时间戳后做基线对比（相似度0.95）',
        })
        return suggestion

    # ── 规则3：含状态关键字 → count检查器
    count_keywords = {
        'FULL':   'OSPF邻居Full状态',
        'Full':   'OSPF邻居Full状态',
        'UP':     '链路Up状态',
        'Up':     '链路Up状态',
        'Active': 'BGP Active状态',
        'ESTABLISHED': 'BGP Established状态',
    }
    for kw, desc in count_keywords.items():
        cnt = text.count(kw)
        if cnt > 0:
            suggestion.update({
                'parser':         'raw',
                'parser_config':  None,
                'checker':        'count',
                'checker_config': {'keyword': kw, 'expected': cnt},
                'reason':         f'检测到关键字 "{kw}"（{desc}）出现 {cnt} 次，建议计数检查',
            })
            return suggestion

    # ── 规则4：单行可提取数字 → regex + threshold
    num_match = _re.search(r'(\d+\.?\d*)\s*(%|Mbps|mbps|ms|GB|MB|KB)', text)
    if num_match:
        val = float(num_match.group(1))
        unit = num_match.group(2)
        pat = _re.escape(text[:num_match.start()]) + r'([\d\.]+)' + _re.escape(unit)
        suggestion.update({
            'parser':         'regex',
            'parser_config':  {'pattern': pat, 'group': 1, 'cast': 'float'},
            'checker':        'threshold',
            'checker_config': {'warning': round(val * 1.3, 2), 'operator': '<='},
            'reason':         f'检测到带单位数值 {val}{unit}，建议正则提取后做阈值检查，警戒值={round(val*1.3,2)}{unit}',
        })
        return suggestion

    # ── 规则5：包含must-have关键字（默认contains）
    must_words = []
    for kw in ['synchronized', 'OK', 'Normal', 'normal', 'success', 'Success']:
        if kw in text:
            must_words.append(kw)
    if must_words:
        suggestion.update({
            'parser':         'raw',
            'parser_config':  None,
            'checker':        'contains',
            'checker_config': {'must_contain': must_words},
            'reason':         f'检测到状态关键字 {must_words}，建议包含检查',
        })
        return suggestion

    # ── 默认：原始基线对比
    return suggestion


# ── Checker 微调工具 ──────────────────────────────────────────
def test_checker_page(request):
    """Checker 微调工具 Web 版：粘贴原始log + 选checker → 立即看结果"""
    from app02.engine.pipeline import PARSERS, CHECKERS, _CUSTOM_CHECKERS
    try:
        import app02.custom_checks  # noqa
    except ImportError:
        pass
    # 检查项预设配置（下拉选择后自动填入 parser/checker/config）
    import json as _json
    check_presets = [
        {
            "label": "【硬件】风扇状态",
            "command": "display fan",
            "parser": "raw", "parser_config": "{}",
            "checker": "custom", "checker_config": _json.dumps({"func": "check_fan"}, ensure_ascii=False),
            "note": "所有风扇Normal→正常"
        },
        {
            "label": "【硬件】电源状态",
            "command": "display power",
            "parser": "raw", "parser_config": "{}",
            "checker": "custom", "checker_config": _json.dumps({"func": "check_power"}, ensure_ascii=False),
            "note": "所有电源Normal→正常"
        },
        {
            "label": "【硬件】温度环境",
            "command": "display environment",
            "parser": "raw", "parser_config": "{}",
            "checker": "custom", "checker_config": _json.dumps({"func": "check_env", "temp_warning": 60}, ensure_ascii=False),
            "note": "温度<阈值(默认60°C)且无Fault"
        },
        {
            "label": "【硬件】单板状态",
            "command": "display device",
            "parser": "raw", "parser_config": "{}",
            "checker": "custom", "checker_config": _json.dumps({"func": "check_device"}, ensure_ascii=False),
            "note": "所有单板Normal→正常"
        },
        {
            "label": "【接口】接口状态（基线对比）",
            "command": "display interface brief",
            "parser": "raw", "parser_config": "{}",
            "checker": "baseline", "checker_config": _json.dumps({"similarity": 1.0}, ensure_ascii=False),
            "note": "A类基线全量对比，清晰展示UP/DOWN变化"
        },
        {
            "label": "【接口】链路聚合",
            "command": "display link-aggregation summary",
            "parser": "raw", "parser_config": "{}",
            "checker": "custom", "checker_config": _json.dumps({"func": "check_agg"}, ensure_ascii=False),
            "note": "无Unselected端口→正常"
        },
        {
            "label": "【二层】STP状态",
            "command": "display stp brief",
            "parser": "raw", "parser_config": "{}",
            "checker": "custom", "checker_config": _json.dumps({"func": "check_stp", "root_expected": "非根桥"}, ensure_ascii=False),
            "note": "非根桥、端口FWD、无阻塞"
        },
        {
            "label": "【二层】VLAN清单",
            "command": "display vlan brief",
            "parser": "raw", "parser_config": "{}",
            "checker": "custom", "checker_config": _json.dumps({"func": "check_vlan", "expected_vlans": []}, ensure_ascii=False),
            "note": "实际VLAN集合=期望(需填expected_vlans)"
        },
        {
            "label": "【路由】OSPF邻居",
            "command": "display ospf peer",
            "parser": "raw", "parser_config": "{}",
            "checker": "custom", "checker_config": _json.dumps({"func": "check_ospf_peer", "expected_full_count": 0}, ensure_ascii=False),
            "note": "所有邻居Full且数量=期望"
        },
        {
            "label": "【路由】BGP邻居",
            "command": "display bgp peer",
            "parser": "raw", "parser_config": "{}",
            "checker": "custom", "checker_config": _json.dumps({"func": "check_bgp_peer", "expected_established": 0}, ensure_ascii=False),
            "note": "Established数=期望"
        },
        {
            "label": "【路由】路由表（基线对比）",
            "command": "display ip routing-table all-vpn-instance",
            "parser": "raw", "parser_config": "{}",
            "checker": "custom", "checker_config": _json.dumps({"similarity": 1.0}, ensure_ascii=False),
            "note": "A类基线全量对比含所有VPN实例，路由增删一目了然"
        },
        {
            "label": "【高可用】VRRP状态",
            "command": "display vrrp verbose",
            "parser": "raw", "parser_config": "{}",
            "checker": "custom", "checker_config": _json.dumps({"func": "check_vrrp"}, ensure_ascii=False),
            "note": "Master数=期望，无Initialize"
        },
        {
            "label": "【高可用】RBM双机热备",
            "command": "display remote-backup-group status",
            "parser": "raw", "parser_config": "{}",
            "checker": "custom", "checker_config": _json.dumps({"func": "check_rbm"}, ensure_ascii=False),
            "note": "主Active，备Inactive"
        },
        {
            "label": "【高可用】M-LAG状态",
            "command": "display m-lag summary",
            "parser": "raw", "parser_config": "{}",
            "checker": "custom", "checker_config": _json.dumps({"func": "check_mlag"}, ensure_ascii=False),
            "note": "Active/Up，无MAD冲突"
        },
        {
            "label": "【防火墙】安全域成员",
            "command": "display security-zone",
            "parser": "raw", "parser_config": "{}",
            "checker": "custom", "checker_config": _json.dumps({"func": "check_zone", "expected": {}}, ensure_ascii=False),
            "note": "各zone接口=期望(需填map)"
        },
        {
            "label": "【防火墙】安全策略规则",
            "command": "display security-policy ip rule all",
            "parser": "raw", "parser_config": "{}",
            "checker": "custom", "checker_config": _json.dumps({"func": "check_security_policy_zone"}, ensure_ascii=False),
            "note": "策略输出非空即正常"
        },
        {
            "label": "【性能】CPU利用率",
            "command": "display cpu-usage",
            "parser": "regex", "parser_config": _json.dumps({"pattern": "(\\d+)%", "group": 1, "cast": "float"}, ensure_ascii=False),
            "checker": "custom", "checker_config": _json.dumps({"func": "check_cpu", "warning": 80, "operator": "<"}, ensure_ascii=False),
            "note": "各核<80%为正常"
        },
        {
            "label": "【性能】内存利用率",
            "command": "display memory",
            "parser": "raw", "parser_config": "{}",
            "checker": "custom", "checker_config": _json.dumps({"func": "check_memory", "warning": 20, "operator": ">"}, ensure_ascii=False),
            "note": "FreeRatio>20%为正常"
        },
        {
            "label": "【性能】会话表",
            "command": "display session table ipv4",
            "parser": "regex", "parser_config": _json.dumps({"pattern": "([\\d]+)", "group": 1, "cast": "int"}, ensure_ascii=False),
            "checker": "custom", "checker_config": _json.dumps({"func": "check_session", "max_sessions": 500000}, ensure_ascii=False),
            "note": "会话数<max_sessions"
        },
        {
            "label": "【日志】日志缓冲",
            "command": "display logbuffer",
            "parser": "strip_ts", "parser_config": _json.dumps({"patterns": ["%\\w{3}\\s+\\d+\\s+\\d{2}:\\d{2}:\\d{2}:\\d{3}\\s+\\d{4}"]}, ensure_ascii=False),
            "checker": "baseline", "checker_config": _json.dumps({"similarity": 0.95}, ensure_ascii=False),
            "note": "对比基线无新增Error/Critical"
        },
        {
            "label": "【探测】NQA探测结果",
            "command": "display nqa result",
            "parser": "raw", "parser_config": "{}",
            "checker": "custom", "checker_config": _json.dumps({"func": "check_nqa"}, ensure_ascii=False),
            "note": "无failed/Timeout"
        },
        {
            "label": "【探测】Track状态",
            "command": "display track",
            "parser": "raw", "parser_config": "{}",
            "checker": "custom", "checker_config": _json.dumps({"func": "check_track", "expected_tracks": 0}, ensure_ascii=False),
            "note": "所有Track Positive"
        },
        {
            "label": "【探测】ARP冲突记录",
            "command": "display arp user-ip-conflict record",
            "parser": "raw", "parser_config": "{}",
            "checker": "custom", "checker_config": _json.dumps({"func": "check_arp"}, ensure_ascii=False),
            "note": "冲突记录清零"
        },
        {
            "label": "【硬件】光模块收发光",
            "command": "display transceiver diagnosis interface",
            "parser": "raw", "parser_config": "{}",
            "checker": "custom", "checker_config": _json.dumps({"func": "check_transceiver"}, ensure_ascii=False),
            "note": "Temp/Voltage/Bias/RX/TX 越限即异常"
        },
    ]
    # Web 编辑的 checker 覆盖：合并文件版 + DB 版函数名
    try:
        from app02.models import CheckerScript
        _db_names = set(CheckerScript.objects.values_list('name', flat=True))
    except Exception:
        _db_names = set()
    func_list = sorted(set(_CUSTOM_CHECKERS.keys()) | _db_names)
    selected_func = request.GET.get('func', '').strip()

    return render(request, 'test_checker.html', {
        'parsers': list(PARSERS.keys()),
        'checkers': list(CHECKERS.keys()),
        'custom_checkers': sorted(_CUSTOM_CHECKERS.keys()),
        'check_presets': check_presets,
        'func_list': func_list,
        'selected_func': selected_func,
    })


@csrf_exempt
def test_checker_run(request):
    """
    Checker 微调执行接口：POST raw log + parser/checker 配置 → 返回结果 JSON
    """
    import json
    from app02.engine.pipeline import PARSERS, CHECKERS
    from app02.engine.reporter import extract_diff_summary
    try:
        import app02.custom_checks  # noqa
    except ImportError:
        pass

    if request.method != 'POST':
        return JsonResponse({'status': False, 'error': '仅支持POST'})

    raw_text    = request.POST.get('raw_text', '')
    baseline_text = request.POST.get('baseline_text', '')
    parser_name = request.POST.get('parser', 'raw')
    checker_name = request.POST.get('checker', 'baseline')

    if not raw_text.strip():
        return JsonResponse({'status': False, 'error': '请粘贴原始命令输出'})

    def _safe_json(s):
        if not s or not s.strip():
            return {}
        return json.loads(s)

    try:
        p_conf = _safe_json(request.POST.get('parser_config', '{}'))
        c_conf = _safe_json(request.POST.get('checker_config', '{}'))
    except json.JSONDecodeError as e:
        return JsonResponse({'status': False, 'error': f'JSON解析失败: {e}'})

    # Step 1: 解析
    parser_func = PARSERS.get(parser_name)
    if not parser_func:
        return JsonResponse({'status': False, 'error': f'未知解析器: {parser_name}'})

    try:
        parsed = parser_func(raw_text, p_conf)
    except Exception as e:
        return JsonResponse({'status': False, 'error': f'解析失败: {e}', 'parsed': None})

    if parsed is None:
        return JsonResponse({
            'status': True,
            'check_result': 'error',
            'notes': '解析结果为 None（正则未匹配），将记为异常',
            'parsed': str(None),
            'parsed_type': 'NoneType',
        })

    # Step 2: 基线解析
    baseline_parsed = None
    if baseline_text.strip():
        try:
            baseline_parsed = parser_func(baseline_text, p_conf)
        except Exception:
            baseline_parsed = None

    # Step 3: 检查
    checker_func = CHECKERS.get(checker_name)
    if not checker_func:
        return JsonResponse({'status': False, 'error': f'未知检查器: {checker_name}'})

    try:
        is_ok, notes = checker_func(parsed, baseline_parsed, c_conf, {})
    except Exception as e:
        return JsonResponse({
            'status': True,
            'check_result': 'error',
            'notes': f'检查器执行异常: {e}',
            'parsed': str(parsed)[:500],
            'parsed_type': type(parsed).__name__,
        })

    result = {
        'status': True,
        'check_result': 'ok' if is_ok else 'anomaly',
        'notes': notes if not is_ok else '',
        'parsed': str(parsed)[:500],
        'parsed_type': type(parsed).__name__,
        'baseline_parsed': str(baseline_parsed)[:200] if baseline_parsed is not None else None,
    }

    # baseline checker 额外输出 diff
    if checker_name == 'baseline' and baseline_text.strip() and not is_ok:
        curr_sum, base_sum, diff_lines = extract_diff_summary(raw_text, baseline_text)
        result['diff_lines'] = diff_lines[:30]

    return JsonResponse(result)


@csrf_exempt
def checker_script_source(request):
    """返回某 checker 的源码（DB覆盖优先，否则文件版）及版本历史。"""
    import inspect
    from app02.engine.pipeline import _CUSTOM_CHECKERS
    from app02.models import CheckerScript, CheckerScriptVersion
    try:
        import app02.custom_checks  # noqa 确保文件版函数已注册
    except ImportError:
        pass
    name = request.GET.get('func', '').strip()
    if not name:
        return JsonResponse({'status': False, 'error': '缺少 func 参数'})
    try:
        row = CheckerScript.objects.get(name=name)
        versions = [{'version': v.version, 'note': v.note,
                     'created_at': v.created_at.strftime('%Y-%m-%d %H:%M')}
                    for v in row.versions.all()]
        return JsonResponse({'status': True, 'func': name, 'source': row.source,
                             'origin': 'db', 'version': row.version,
                             'enabled': row.enabled, 'versions': versions})
    except CheckerScript.DoesNotExist:
        fn = _CUSTOM_CHECKERS.get(name)
        if not fn:
            return JsonResponse({'status': False, 'error': f'未找到 checker: {name}'})
        try:
            src = inspect.getsource(fn)
        except (OSError, TypeError):
            src = '# 该 checker 为 DB 加载对象，无法显示文件源码'
        return JsonResponse({'status': True, 'func': name, 'source': src,
                             'origin': 'file', 'version': None,
                             'enabled': True, 'versions': []})


@csrf_exempt
def checker_script_save(request):
    """保存 checker 源码（DB覆盖）：自动归档旧版本并热加载到 _CUSTOM_CHECKERS。"""
    import ast
    from app02.engine.pipeline import load_checker_overrides
    from app02.models import CheckerScript, CheckerScriptVersion
    if request.method != 'POST':
        return JsonResponse({'status': False, 'error': '仅支持POST'})
    func = (request.POST.get('func') or '').strip()
    source = request.POST.get('source', '')
    note = request.POST.get('note', '').strip()
    enabled = request.POST.get('enabled', 'true') not in ('0', 'false', 'False')
    if not func:
        return JsonResponse({'status': False, 'error': '缺少 func'})
    if not source.strip():
        return JsonResponse({'status': False, 'error': '源码不能为空'})
    try:
        tree = ast.parse(source)
    except SyntaxError as e:
        return JsonResponse({'status': False, 'error': f'语法错误: {e}'})
    defined = [n.name for n in ast.walk(tree) if isinstance(n, ast.FunctionDef)]
    if func not in defined:
        return JsonResponse({'status': False, 'error':
            f'源码需定义 def {func}(parsed, baseline, config, extra)，当前定义: {defined or "无"}'})
    row, created = CheckerScript.objects.get_or_create(
        name=func, defaults={'source': source, 'version': 1, 'enabled': enabled, 'note': note})
    if not created:
        CheckerScriptVersion.objects.create(script=row, version=row.version,
                                            source=row.source, note=row.note or '自动归档')
        row.source = source
        row.version = row.version + 1
        row.enabled = enabled
        row.note = note
        row.save()
    try:
        load_checker_overrides()
    except Exception as e:
        return JsonResponse({'status': True, 'saved': True,
                             'warning': f'已保存但热加载失败(重启后生效): {e}'})
    return JsonResponse({'status': True, 'saved': True, 'version': row.version,
                         'message': f'已保存并热加载 {func} (v{row.version})'})


@csrf_exempt
def checker_script_rollback(request):
    """回滚 checker 到指定历史版本。"""
    from app02.engine.pipeline import load_checker_overrides
    from app02.models import CheckerScript, CheckerScriptVersion
    if request.method != 'POST':
        return JsonResponse({'status': False, 'error': '仅支持POST'})
    func = (request.POST.get('func') or '').strip()
    try:
        ver = int(request.POST.get('version', '0'))
    except ValueError:
        return JsonResponse({'status': False, 'error': 'version 无效'})
    if not func:
        return JsonResponse({'status': False, 'error': '缺少 func'})
    try:
        row = CheckerScript.objects.get(name=func)
        target = row.versions.get(version=ver)
    except (CheckerScript.DoesNotExist, CheckerScriptVersion.DoesNotExist) as e:
        return JsonResponse({'status': False, 'error': f'未找到: {e}'})
    CheckerScriptVersion.objects.create(script=row, version=row.version,
                                        source=row.source, note='回滚前自动归档')
    row.source = target.source
    row.version = row.version + 1
    row.save()
    try:
        load_checker_overrides()
    except Exception as e:
        return JsonResponse({'status': True, 'warning': f'已保存但热加载失败: {e}'})
    return JsonResponse({'status': True, 'message': f'已回滚 {func} 至历史 v{ver} (新版本 v{row.version})'})


def new_import_page(request):
    """AI辅助批量导入巡检项首页"""
    groups = DeviceGroup.objects.all()
    return render(request, 'new_import.html', {'groups': groups})


@csrf_exempt
def new_ai_parse(request):
    """
    AI解析接口：接收命令名+输出文本，返回建议配置。
    POST: name, command, output
    """
    if request.method != 'POST':
        return JsonResponse({'status': False, 'error': '仅支持POST'})

    name    = request.POST.get('name', '').strip()
    command = request.POST.get('command', '').strip()
    output  = request.POST.get('output', '').strip()

    if not command or not output:
        return JsonResponse({'status': False, 'error': '命令和输出内容不能为空'})

    suggestion = _ai_analyze_output(command, output)
    return JsonResponse({'status': True, 'data': suggestion})


@csrf_exempt
def new_ai_batch_import(request):
    """
    批量导入确认接口：接收多条巡检项JSON，批量写入CheckItem表。
    POST body: items (JSON array)
    """
    if request.method != 'POST':
        return JsonResponse({'status': False, 'error': '仅支持POST'})

    try:
        body  = json.loads(request.body)
        items = body.get('items', [])
    except (json.JSONDecodeError, AttributeError):
        return JsonResponse({'status': False, 'error': 'JSON解析失败'})

    created = []
    errors  = []
    for item in items:
        name    = str(item.get('name', '')).strip()
        command = str(item.get('command', '')).strip()
        if not name or not command:
            errors.append(f'跳过：名称或命令为空 {item}')
            continue
        try:
            obj = CheckItem.objects.create(
                name           = name,
                command        = command,
                parser         = item.get('parser', 'raw'),
                parser_config  = item.get('parser_config'),
                checker        = item.get('checker', 'baseline'),
                checker_config = item.get('checker_config'),
                error_note     = item.get('error_note', '请检查'),
                timeout        = int(item.get('timeout', 30)),
                enabled        = True,
            )
            # 可选：绑定到分组
            group_id = item.get('group_id')
            if group_id:
                grp = DeviceGroup.objects.filter(id=group_id).first()
                if grp:
                    grp.check_items.add(obj)
            created.append({'id': obj.id, 'name': obj.name})
        except Exception as e:
            errors.append(f'{name}: {e}')

    return JsonResponse({
        'status':  True,
        'created': len(created),
        'errors':  errors,
        'items':   created,
    })


def dashboard(request):
    """仪表盘：汇总设备/巡检概况（阶段A外壳美化，数据来自现有模型）。"""
    total_devices = NewDevice.objects.count()
    total_checksets = CheckSet.objects.count()
    total_checkitems = CheckItem.objects.count()
    pending_anomaly = AnomalyRecord.objects.filter(confirm=False).count()

    latest = XunjianRecord.objects.first()  # Meta.ordering = ['-time']
    ok = anomaly = failed = 0
    healthy_pct = dash_ok = pct_ok = pct_anomaly = pct_failed = 0
    latest_time = ''
    if latest:
        dc = latest.device_count or 0
        ok = latest.ok_devices or 0
        anomaly = latest.anomaly_devices or 0
        failed = latest.failed_devices or 0
        latest_time = latest.time
        if dc:
            healthy_pct = round(ok / dc * 100)
            dash_ok = round(healthy_pct / 100 * 289)
            pct_ok = round(ok / dc * 100)
            pct_anomaly = round(anomaly / dc * 100)
            pct_failed = round(failed / dc * 100)

    recent = []
    for r in XunjianRecord.objects.all()[:5]:
        dc = r.device_count or 0
        r_ok = r.ok_devices or 0
        pct = round(r_ok / dc * 100) if dc else 0
        res = r.result or ''
        if res == '正常':
            badge = 'ok'
        elif '异常' in res:
            badge = 'warn'
        else:
            badge = 'fail'
        recent.append({
            'time': r.time, 'operator': r.operator, 'device_count': dc,
            'ok': r_ok, 'pct': pct, 'result': res, 'badge': badge,
        })

    return render(request, 'dashboard.html', {
        'total_devices': total_devices,
        'total_checksets': total_checksets,
        'total_checkitems': total_checkitems,
        'pending_anomaly': pending_anomaly,
        'ok': ok, 'anomaly': anomaly, 'failed': failed,
        'healthy_pct': healthy_pct, 'dash_ok': dash_ok,
        'pct_ok': pct_ok, 'pct_anomaly': pct_anomaly, 'pct_failed': pct_failed,
        'latest_time': latest_time, 'recent': recent,
    })


def task_center(request):
    """任务中心：列出巡检任务（阶段A后台线程 + 任务表），含进度与状态。"""
    tasks = XunjianTask.objects.all()[:50]
    data = []
    for t in tasks:
        dc = t.device_count or 0
        done = t.done or 0
        pct = round(done / dc * 100) if dc else 0
        status = t.status
        if status == 'done':
            badge = 'ok'
        elif status in ('queued', 'running'):
            badge = 'run'
        elif status == 'partial':
            badge = 'warn'
        else:
            badge = 'fail'
        data.append({
            'id': t.id,
            'status': status,
            'status_cn': t.get_status_display(),
            'operator': t.operator,
            'device_count': dc,
            'done': done,
            'pct': pct,
            'ok': t.ok_devices,
            'anomaly': t.anomaly_devices,
            'failed': t.failed_devices,
            'badge': badge,
            'created_at': t.created_at.strftime('%Y-%m-%d %H:%M'),
            'xunjian_time': t.xunjian_time,
            'has_failed': bool(t.failed_device_list),
        })
    return render(request, 'task_center.html', {'tasks': data})


def task_detail(request, task_id):
    """任务详情页：展示实时进度（前端轮询 /task/<id>/detail/）。"""
    try:
        task = XunjianTask.objects.get(id=task_id)
    except XunjianTask.DoesNotExist:
        return redirect('/task/center/')
    return render(request, 'task_detail.html', {
        'task_id':      task.id,
        'status':       task.status,
        'status_cn':    task.get_status_display(),
        'operator':     task.operator,
        'device_count': task.device_count,
        'xunjian_time': task.xunjian_time,
        'has_failed':   bool(task.failed_device_list),
        'created_at':   task.created_at.strftime('%Y-%m-%d %H:%M'),
    })


def task_detail_json(request, task_id):
    """进度轮询接口（前端每 1.5s 调用一次）。"""
    try:
        task = XunjianTask.objects.get(id=task_id)
    except XunjianTask.DoesNotExist:
        return JsonResponse({'status': False, 'error': '任务不存在'}, status=404)
    dc = task.device_count or 0
    done = task.done or 0
    pct = round(done / dc * 100) if dc else 0
    status = task.status
    if status == 'done':
        badge = 'ok'
    elif status in ('queued', 'running'):
        badge = 'run'
    elif status == 'partial':
        badge = 'warn'
    else:
        badge = 'fail'
    return JsonResponse({
        'status': True,
        'task': {
            'id':            task.id,
            'status':        status,
            'status_cn':     task.get_status_display(),
            'badge':         badge,
            'device_count':  dc,
            'done':          done,
            'pct':           pct,
            'ok_devices':    task.ok_devices,
            'anomaly_devices': task.anomaly_devices,
            'failed_devices':  task.failed_devices,
            'has_failed':    bool(task.failed_device_list),
            'xunjian_time':  task.xunjian_time,
            'result':        task.result,
            'error':         task.error,
        }
    })


@csrf_exempt
def task_resume(request, task_id):
    """续跑失败设备：仅重新巡检原任务的失败设备。"""
    if request.method != 'POST':
        return JsonResponse({'status': False, 'error': '仅支持POST'}, status=405)
    try:
        old = XunjianTask.objects.get(id=task_id)
    except XunjianTask.DoesNotExist:
        return JsonResponse({'status': False, 'error': '任务不存在'}, status=404)
    if not old.failed_device_list:
        return JsonResponse({'status': False, 'error': '无失败设备可续跑'}, status=400)

    names = [n.strip() for n in old.failed_device_list.split(',') if n.strip()]
    devs = NewDevice.objects.filter(name__in=names, enabled=True)
    ids = list(devs.values_list('id', flat=True))
    if not ids:
        return JsonResponse({'status': False, 'error': '找不到对应的失败设备'}, status=400)

    operator = request.session.get('info', {}).get('name', '未知')
    new_task = XunjianTask.objects.create(
        status='queued', operator=operator,
        checkset=old.checkset, device_count=len(ids),
    )

    def _worker():
        try:
            _run_xunjian_new(operator=operator, device_ids=ids, task_id=new_task.id)
        except Exception as e:
            close_old_connections()
            XunjianTask.objects.filter(id=new_task.id).update(
                status='failed', error=str(e)[:500],
            )
            _new_logger.error(f'续跑失败(task={new_task.id}): {e}')
        finally:
            close_old_connections()

    threading.Thread(target=_worker, daemon=True).start()
    return JsonResponse({'status': True, 'task_id': new_task.id})




# ── 验收报告 ─────────────────────────────────────────────────
def acceptance_report(request):
    """验收报告：基于数据库巡检结果生成 HTML 报告"""
    from django.http import HttpResponse
    from app02.models import (
        XunjianRecord, CheckResult, AnomalyRecord, NewDevice
    )
    from app02.engine.reporter import render_acceptance_report

    xunjian_time = request.GET.get('time', '')
    if not xunjian_time:
        # 取最新一条记录
        record = XunjianRecord.objects.order_by('-time').first()
    else:
        record = XunjianRecord.objects.filter(time=xunjian_time).first()

    if not record:
        return HttpResponse('<h3>未找到巡检记录</h3><p>请先执行一次巡检。</p>')

    # 获取此次巡检的所有结果
    check_results = CheckResult.objects.filter(time=record.time)
    anomalies = AnomalyRecord.objects.filter(time=record.time)
    device_names = check_results.values_list('device', flat=True).distinct()
    devices = NewDevice.objects.filter(name__in=device_names)

    sites = sorted({d.site for d in devices if d.site})
    site_label = sites[0] if len(sites) == 1 else ('/'.join(sites) if sites else '网络巡检')

    html = render_acceptance_report(record, anomalies, check_results, devices, site_label=site_label)
    return HttpResponse(html)


# ═══════════════════════════════════════════════════════
# 阶段 C：设备发现 + 配置合规
# ═══════════════════════════════════════════════════════
@csrf_exempt
def stage_cd_run(request):
    """运行设备发现 / 配置合规（POST）。完成后重定向回页面（结果已写入 DB）。"""
    if request.method != 'POST':
        return JsonResponse({'status': False, 'error': '仅支持POST'})
    from django.shortcuts import redirect
    from app02.engine.stage_c import run_discovery, run_compliance
    site = request.POST.get('site', '') or None
    kind = request.POST.get('kind', 'all')  # all / discovery / compliance
    try:
        if kind in ('all', 'discovery'):
            run_discovery(site=site)
        if kind in ('all', 'compliance'):
            run_compliance(site=site)
    except Exception as e:
        return JsonResponse({'status': False, 'error': str(e)})
    # 普通表单提交 → 重定向回页面；AJAX → 返回 JSON
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({'status': True})
    return redirect('/new/stagecd/' + (f'?site={site}' if site else ''))


def stage_cd_page(request):
    from app02.models import (DiscoveryRecord, ComplianceResult,
                              CompliancePolicy, NewDevice)
    site = request.GET.get('site', '')
    latest_disc = DiscoveryRecord.objects.order_by('-created_at', '-id').first()
    latest_comp = ComplianceResult.objects.order_by('-created_at', '-id').first()
    disc_time = latest_disc.time if latest_disc else ''
    comp_time = latest_comp.time if latest_comp else ''

    unknowns = (DiscoveryRecord.objects.filter(time=disc_time, is_known=False)
                if disc_time else [])
    known_devs = set(DiscoveryRecord.objects.filter(time=disc_time)
                     .values_list('device', flat=True)) if disc_time else set()
    no_lldp = (NewDevice.objects.filter(enabled=True).exclude(name__in=known_devs)
               if disc_time else [])
    violations = (ComplianceResult.objects.filter(time=comp_time, passed=False)
                  if comp_time else [])
    policies = CompliancePolicy.objects.all()
    return render(request, 'stage_cd.html', {
        'site': site, 'disc_time': disc_time, 'comp_time': comp_time,
        'unknowns': unknowns, 'no_lldp': no_lldp, 'violations': violations,
        'policies': policies,
    })


# ═══════════════════════════════════════════════════════
# 阶段 D：趋势图 + 验收报告导出
# ═══════════════════════════════════════════════════════
METRIC_DEFS = {
    'cpu':    {'command': 'display cpu-usage', 'regex': r'(\d+)%\s+in last 5 seconds',
               'unit': '%', 'label': 'CPU使用率'},
    'memory': {'command': 'display memory', 'regex': r'Mem:[\s\d]+\s+([\d.]+)%',
               'unit': '%', 'label': '内存空闲率(FreeRatio)'},
    'temp':   {'command': 'display environment', 'regex': r'(\d+)',
               'unit': '°C', 'label': '环境温度'},
}


def _trend_series(device, metric):
    import re
    from app02.models import CheckResult
    d = METRIC_DEFS.get(metric)
    if not d or not device:
        return []
    rows = (CheckResult.objects
            .filter(device=device, command=d['command'])
            .exclude(created_at__isnull=True)
            .order_by('created_at'))
    series = []
    for r in rows:
        if not r.result:
            continue
        m = re.search(d['regex'], r.result)
        if not m:
            continue
        series.append({'t': r.created_at.strftime('%Y-%m-%d %H:%M'),
                       'v': float(m.group(1))})
    return series


def _trend_chart_svg(series, label, unit):
    """服务端生成 SVG 折线图（无外网/CDN 依赖）。"""
    W, H, pad = 680, 260, 44
    if not series:
        return ('<svg xmlns="http://www.w3.org/2000/svg" width="%d" height="%d">'
                '<rect width="100%%" height="100%%" fill="#fff"/>'
                '<text x="24" y="%d" fill="#888" font-size="14">暂无趋势数据（请先积累多次巡检）</text>'
                '</svg>' % (W, H, H // 2))
    vals = [p['v'] for p in series]
    vmin, vmax = min(vals), max(vals)
    if vmax == vmin:
        vmax, vmin = vmin + 1, max(0, vmin - 1)
    n = len(series)

    def X(i):
        return pad + (W - 2 * pad) * (i / (n - 1)) if n > 1 else W / 2

    def Y(v):
        return H - pad - (H - 2 * pad) * ((v - vmin) / (vmax - vmin))

    grid = ''
    for g in range(4):
        gy = pad + (H - 2 * pad) * g / 3
        gv = vmax - (vmax - vmin) * g / 3
        grid += (f'<line x1="{pad}" y1="{gy:.1f}" x2="{W-pad}" y2="{gy:.1f}" '
                 f'stroke="#eee"/><text x="4" y="{gy+4:.1f}" fill="#999" font-size="11">{gv:.1f}</text>')
    pts = ' '.join(f'{X(i):.1f},{Y(v):.1f}' for i, v in enumerate(vals))
    last = series[-1]
    return (f'<svg xmlns="http://www.w3.org/2000/svg" width="{W}" height="{H}">'
            f'<rect width="100%" height="100%" fill="#fff"/>{grid}'
            f'<polyline fill="none" stroke="#0F6E56" stroke-width="2" points="{pts}"/>'
            f'<circle cx="{X(n-1):.1f}" cy="{Y(vals[-1]):.1f}" r="4" fill="#0F6E56"/>'
            f'<text x="{W-pad}" y="{Y(vals[-1])-10:.1f}" fill="#0F6E56" font-size="12" '
            f'text-anchor="end">{last["v"]:.1f}{unit}</text>'
            f'<text x="{pad}" y="16" fill="#333" font-size="13">{label} 趋势（{n} 个采样点）</text>'
            f'</svg>')


def trend_page(request):
    from app02.models import NewDevice
    device = request.GET.get('device', '')
    metric = request.GET.get('metric', 'cpu')
    devices = list(NewDevice.objects.filter(enabled=True).values_list('name', flat=True))
    if device and device not in devices:
        devices.insert(0, device)
    series = _trend_series(device, metric) if device else []
    svg = _trend_chart_svg(series, METRIC_DEFS.get(metric, {}).get('label', ''),
                           METRIC_DEFS.get(metric, {}).get('unit', ''))
    return render(request, 'trend.html', {
        'devices': devices, 'device': device, 'metric': metric,
        'metrics': METRIC_DEFS, 'svg': mark_safe(svg), 'series': series,
    })


def acceptance_report_export(request):
    """验收报告导出为 Excel（openpyxl）。"""
    from django.http import HttpResponse
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from app02.models import (XunjianRecord, CheckResult, AnomalyRecord, NewDevice)

    xunjian_time = request.GET.get('time', '') or request.GET.get('xunjian_time', '')
    record = (XunjianRecord.objects.filter(time=xunjian_time).first()
              if xunjian_time else XunjianRecord.objects.order_by('-time').first())
    if not record:
        return HttpResponse('未找到巡检记录，请先执行一次巡检。', status=404)

    check_results = CheckResult.objects.filter(time=record.time)
    anomalies = list(AnomalyRecord.objects.filter(time=record.time))
    device_names = list(check_results.values_list('device', flat=True).distinct())
    devices = NewDevice.objects.filter(name__in=device_names)
    sites = sorted({d.site for d in devices if d.site})
    site_label = sites[0] if len(sites) == 1 else ('/'.join(sites) if sites else '网络巡检')
    dev_status = {d.name: '异常' for d in anomalies}
    dev_map = {d.name: d for d in devices}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '巡检汇总'
    head_fill = PatternFill('solid', fgColor='0F6E56')
    head_font = Font(color='FFFFFF', bold=True)
    ws.append([f'{site_label}网络巡检验收报告'])
    ws['A1'].font = Font(size=14, bold=True)
    ws.append(['巡检时间', record.time, '总结果', record.result])
    ws.append(['设备数', record.device_count, '异常台数', record.anomaly_devices,
               '失败台数', record.failed_devices])
    ws.append([])
    ws.append(['设备', '命令', '严重级别', '异常说明', '当前值', '基线值'])
    for c in range(1, 7):
        ws.cell(row=5, column=c).fill = head_fill
        ws.cell(row=5, column=c).font = head_font
    for a in anomalies:
        ws.append([a.device, a.command, a.severity, a.notes,
                   (a.current_val or '')[:200], (a.baseline_val or '')[:200]])

    ws2 = wb.create_sheet('设备明细')
    ws2.append(['设备', 'IP', '角色', '站点', '状态'])
    for c in range(1, 6):
        ws2.cell(row=1, column=c).fill = head_fill
        ws2.cell(row=1, column=c).font = head_font
    for name in device_names:
        d = dev_map.get(name)
        ws2.append([name, d.ip if d else '', d.role if d else '',
                    d.site if d else '', dev_status.get(name, '正常')])

    resp = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename=acceptance_{record.time}.xlsx'
    wb.save(resp)
    return resp


# ════════════════════════════════════════════════
# 整体巡检报告（项目设备整体态势，独立于单次巡检）
# ════════════════════════════════════════════════
def _fleet_stats(site_filter=''):
    """
    汇总「项目整体态势」数据，供整体巡检报告页与导出复用。
    数据来源：
      - NewDevice(启用)        设备资产台账（总数 / 站点 / 角色）
      - CmdbDevice(最新同步)    资源健康（CPU/内存阈值预警）
      - AnomalyRecord(未确认)   待处理异常（累计未确认 = 当前待办）
      - XunjianRecord           最近巡检概览
    """
    from app02.models import CmdbDevice
    from collections import defaultdict

    site_filter = (site_filter or '').strip()
    in_site = lambda s: (not site_filter) or (s == site_filter)

    # 1) 设备资产台账（启用设备）
    dev_rows = list(NewDevice.objects.filter(enabled=True)
                    .values('name', 'ip', 'site', 'role', 'group__name'))
    meta = {d['name']: d for d in dev_rows}
    total = len(dev_rows)

    # 2) CMDB 资源健康（最新同步快照）
    cmdb_all = list(CmdbDevice.objects.all().values(
        'name', 'site', 'role', 'cpu_5s', 'mem_free_ratio', 'uptime_days',
        'model', 'serial', 'last_sync'))
    warn_list = []
    for c in cmdb_all:
        if not in_site(c['site']):
            continue
        reasons = []
        if c['cpu_5s'] is not None and c['cpu_5s'] >= 80:
            reasons.append('CPU高负载')
        if c['mem_free_ratio'] is not None and c['mem_free_ratio'] < 20:
            reasons.append('内存紧张')
        if reasons:
            warn_list.append({
                'name': c['name'], 'site': c['site'], 'role': c['role'] or '-',
                'cpu': c['cpu_5s'], 'mem': c['mem_free_ratio'],
                'uptime': c['uptime_days'], 'reasons': '、'.join(reasons),
                'model': c['model'], 'serial': c['serial'],
            })
    warn_count = len(warn_list)

    # 3) 站点 / 角色 设备分布
    site_dev = defaultdict(int)
    role_dev = defaultdict(int)
    for d in dev_rows:
        if not in_site(d['site']):
            continue
        site_dev[d['site'] or '未标注'] += 1
        role_dev[d['role'] or '未分类'] += 1
    site_max = max(site_dev.values(), default=1)
    role_max = max(role_dev.values(), default=1)
    site_dist = [{'name': k, 'count': v, 'pct': round(v / site_max * 100)}
                 for k, v in sorted(site_dev.items())]
    role_dist = [{'name': k, 'count': v, 'pct': round(v / role_max * 100)}
                 for k, v in sorted(role_dev.items())]

    cmdb_in_site = sum(1 for c in cmdb_all if in_site(c['site']))
    onsite_total = sum(site_dev.values())
    synced_ok = cmdb_in_site - warn_count
    not_synced = onsite_total - cmdb_in_site

    # 4) 待处理异常（累计未确认）
    anoms = list(AnomalyRecord.objects.filter(confirm=False))
    sev_counts = {'P0': 0, 'P1': 0, 'P2': 0}
    site_anom = defaultdict(int)
    role_anom = defaultdict(int)
    dev_anom = defaultdict(list)      # device -> [(command, notes, severity), ...]
    for a in anoms:
        m = meta.get(a.device)
        if site_filter and (not m or m['site'] != site_filter):
            continue
        sev = a.severity or 'P2'
        sev_counts[sev] = sev_counts.get(sev, 0) + 1
        if m:
            site_anom[m['site'] or '未标注'] += 1
            role_anom[m['role'] or '未分类'] += 1
        dev_anom[a.device].append((a.command, a.notes or '异常', sev))

    p0 = sev_counts.get('P0', 0)
    p1 = sev_counts.get('P1', 0)
    p2 = sev_counts.get('P2', 0)
    total_anom = p0 + p1 + p2
    high = p0 + p1
    sev_order = {'P0': 0, 'P1': 1, 'P2': 2}
    sev_dist = []
    for sev, color in (('P0', '#d14343'), ('P1', '#b7791f'), ('P2', '#9aa3af')):
        c = sev_counts.get(sev, 0)
        sev_dist.append({'sev': sev, 'count': c,
                         'pct': (round(c / total_anom * 100) if total_anom else 0),
                         'color': color})
    site_anom_dist = [{'name': k, 'count': v,
                       'pct': round(v / max(site_anom.values(), default=1) * 100)}
                      for k, v in sorted(site_anom.items())]
    role_anom_dist = [{'name': k, 'count': v,
                       'pct': round(v / max(role_anom.values(), default=1) * 100)}
                      for k, v in sorted(role_anom.items())]

    # 5) 问题清单（按 命令+说明 聚合）
    grouped = {}
    for dev, alist in dev_anom.items():
        for cmd, notes, sev in alist:
            key = (cmd, notes)
            g = grouped.get(key)
            if g is None:
                g = grouped[key] = {'command': cmd, 'notes': notes,
                                    'sev': sev, 'devices': set()}
            g['devices'].add(dev)
            if sev_order.get(sev, 2) < sev_order.get(g['sev'], 2):
                g['sev'] = sev
    issues = [{'command': g['command'], 'notes': g['notes'], 'severity': g['sev'],
               'device_count': len(g['devices']), 'devices': sorted(g['devices'])}
              for g in grouped.values()]
    issues.sort(key=lambda x: sev_order.get(x['severity'], 2))

    # 异常 Top 设备
    top_devs = sorted(dev_anom.items(), key=lambda kv: len(kv[1]), reverse=True)[:12]
    top_devices = []
    for dev, alist in top_devs:
        sev = max((s for _, _, s in alist), key=lambda s: sev_order.get(s, 2))
        top_devices.append({'name': dev, 'count': len(alist), 'sev': sev,
                            'meta': meta.get(dev)})

    # 6) 最近巡检
    recent = []
    for r in XunjianRecord.objects.all()[:6]:
        dc = r.device_count or 0
        recent.append({'time': r.time, 'operator': r.operator, 'device_count': dc,
                       'ok': r.ok_devices or 0, 'anomaly': r.anomaly_devices or 0,
                       'failed': r.failed_devices or 0, 'result': r.result or ''})

    # 7) 数据时间锚点
    latest_sync = (CmdbDevice.objects.exclude(last_sync=None)
                   .order_by('-last_sync').values_list('last_sync', flat=True).first())
    latest_run = XunjianRecord.objects.first()
    latest_run_time = latest_run.time if latest_run else ''

    healthy_pct = round(synced_ok / onsite_total * 100) if onsite_total else 0

    # ════════════════════════════════════════════════════════════
    # 分析层：把"数据罗列"升级为"整体总结分析"
    #   - 整体健康评级（覆盖度 + 已巡检设备健康加权）
    #   - 跨轮次趋势（按 XunjianRecord 展开，叠加每轮严重级别）
    #   - 故障模式归因（按巡检项聚合，标记系统性/局部）
    #   - 风险集中度（按角色 / 站点异常率）
    #   - 覆盖缺口（启用但未建台账/未巡检设备）
    #   - 自动建议（由上述指标推导）
    # ════════════════════════════════════════════════════════════
    open_dev_n = len(dev_anom)                      # 存在未确认异常的设备数（本站点）
    inspected = cmdb_in_site                        # 已建台账(已巡检)设备数

    # 覆盖度
    coverage_pct = round(inspected / onsite_total * 100) if onsite_total else 0
    synced_names = {c['name'] for c in cmdb_all if in_site(c['site'])}
    enabled_names = {d['name'] for d in dev_rows if in_site(d['site'])}
    never_synced = sorted(enabled_names - synced_names)

    # 已巡检设备健康分（按严重级别加权的异常负担）
    if inspected:
        W = p0 * 4 + p1 * 2 + p2 * 0.5
        burden_per_dev = W / inspected
        health_inspected = max(0.0, 100 - min(75.0, burden_per_dev * 12))
    else:
        health_inspected = 0.0
    blended = health_inspected * 0.6 + coverage_pct * 0.4
    # 存在未处置 P0 高危时，评级封顶为 B（不允许评优秀）
    if p0 > 0 and blended >= 85:
        blended = min(blended, 84.0)
    health_score = round(health_inspected)
    if blended >= 85:   grade, grade_label, grade_color = 'A', '优秀', '#0F6E56'
    elif blended >= 70: grade, grade_label, grade_color = 'B', '良好', '#3a8f6b'
    elif blended >= 55: grade, grade_label, grade_color = 'C', '一般', '#b7791f'
    elif blended >= 40: grade, grade_label, grade_color = 'D', '较差', '#d9822b'
    else:               grade, grade_label, grade_color = 'E', '高危', '#d14343'

    # 趋势：按巡检轮次展开，叠加每轮严重级别分布
    runs_chrono = list(reversed(list(XunjianRecord.objects.all()[:12])))
    trend = []
    prev_anom = None
    for r in runs_chrono:
        rt = r.time
        sevs = {'P0': 0, 'P1': 0, 'P2': 0}
        for a in AnomalyRecord.objects.filter(time=rt):
            sevs[a.severity or 'P2'] = sevs.get(a.severity or 'P2', 0) + 1
        dc = r.device_count or 0
        anom = r.anomaly_devices or 0
        rate = round(anom / dc * 100, 1) if dc else 0
        trend.append({
            'time': rt, 'device_count': dc, 'ok': r.ok_devices or 0,
            'anomaly': anom, 'failed': r.failed_devices or 0,
            'p0': sevs['P0'], 'p1': sevs['P1'], 'p2': sevs['P2'],
            'rate': rate,
            'delta': (anom - prev_anom) if prev_anom is not None else None,
        })
        prev_anom = anom
    overall_delta = (trend[-1]['anomaly'] - trend[0]['anomaly']) if len(trend) >= 2 else None

    # 故障模式归因：按巡检项(命令)聚合未确认异常，标记系统性/局部
    fm = {}
    for a in anoms:
        m = meta.get(a.device)
        if site_filter and (not m or m['site'] != site_filter):
            continue
        key = a.command
        g = fm.get(key)
        if g is None:
            g = fm[key] = {'command': key, 'count': 0,
                           'sev_counts': {'P0': 0, 'P1': 0, 'P2': 0},
                           'devices': set(), 'sev': 'P2'}
        g['count'] += 1
        g['devices'].add(a.device)
        sev = a.severity or 'P2'
        g['sev_counts'][sev] += 1
        if sev_order.get(sev, 2) < sev_order.get(g['sev'], 2):
            g['sev'] = sev
    failure_modes = [{'command': g['command'], 'count': g['count'],
                      'severity': g['sev'], 'device_count': len(g['devices']),
                      'sev_counts': g['sev_counts'], 'systemic': len(g['devices']) >= 3}
                     for g in fm.values()]
    failure_modes.sort(key=lambda x: (-x['count'], sev_order.get(x['severity'], 2)))

    # 风险集中度：按角色 / 站点 的异常设备率（有异常的设备数 ÷ 该范围设备数，0~100%）
    role_dev_anom = defaultdict(set)
    site_dev_anom = defaultdict(set)
    for dev in dev_anom:
        m = meta.get(dev)
        if not m:
            continue
        role_dev_anom[m.get('role') or '未分类'].add(dev)
        site_dev_anom[m.get('site') or '未标注'].add(dev)
    risk_by_role = []
    for role, dn in role_dev.items():
        an_dev = len(role_dev_anom.get(role, set()))
        risk_by_role.append({'role': role, 'devices': dn, 'anom': an_dev,
                             'rate': round(an_dev / dn * 100, 1) if dn else 0})
    risk_by_role.sort(key=lambda x: -x['rate'])
    risk_by_site = []
    for st, dn in site_dev.items():
        an_dev = len(site_dev_anom.get(st, set()))
        risk_by_site.append({'site': st, 'devices': dn, 'anom': an_dev,
                             'rate': round(an_dev / dn * 100, 1) if dn else 0})
    risk_by_site.sort(key=lambda x: -x['rate'])

    # 自动建议（由上述指标推导）
    recommendations = []
    if coverage_pct < 95:
        recommendations.append(
            f'{not_synced} 台启用设备尚未纳入巡检/台账（覆盖度 {coverage_pct}%），'
            f'整体可见性不足，建议尽快补巡检。')
    if p0 > 0:
        recommendations.append(f'存在 {p0} 条 P0 高危异常，需立即处置。')
    if failure_modes:
        top = failure_modes[0]
        recommendations.append(
            f'最高频故障项「{top["command"]}」影响 {top["device_count"]} 台设备'
            f'（{"系统性" if top["systemic"] else "局部"}），建议优先排查根因。')
    if warn_count > 0:
        recommendations.append(
            f'{warn_count} 台设备资源水位预警（CPU/内存），关注容量趋势。')
    if p2 > 20:
        recommendations.append(
            f'累计 P2 未确认 {p2} 条，建议批量确认以释放待办。')
    if len(trend) >= 2:
        if trend[-1]['anomaly'] > trend[-2]['anomaly']:
            recommendations.append(
                f'近一次巡检异常数上升（{trend[-2]["anomaly"]}→{trend[-1]["anomaly"]}），'
                f'态势转差，需重点跟进。')
        else:
            recommendations.append(
                f'近一次巡检异常数下降（{trend[-2]["anomaly"]}→{trend[-1]["anomaly"]}），'
                f'态势向好，保持。')

    narrative = (f'整体健康评级 {grade}（{grade_label}），综合得分 {round(blended, 1)} 分。'
                 f'已巡检/建台账 {inspected} 台（覆盖度 {coverage_pct}%）；'
                 f'当前待处理异常 {total_anom} 项（P0 {p0} / P1 {p1} / P2 {p2}），'
                 f'涉及设备 {open_dev_n} 台；资源水位预警 {warn_count} 台。')
    if not_synced:
        narrative += f'另有 {not_synced} 台启用设备尚未巡检，整体可见性受限。'

    return {
        'site_filter': site_filter,
        'total': total,
        'cmdb_count': cmdb_in_site,
        'warn_count': warn_count,
        'warn_list': warn_list,
        'site_dist': site_dist,
        'role_dist': role_dist,
        'synced_ok': synced_ok,
        'not_synced': not_synced,
        'onsite_total': onsite_total,
        'healthy_pct': healthy_pct,
        'p0': p0, 'p1': p1, 'p2': p2,
        'total_anom': total_anom, 'high': high,
        'sev_dist': sev_dist,
        'site_anom_dist': site_anom_dist,
        'role_anom_dist': role_anom_dist,
        'issues': issues,
        'top_devices': top_devices,
        'recent': recent,
        'latest_sync': latest_sync,
        'latest_run_time': latest_run_time,
        # —— 分析层 ——
        'grade': grade, 'grade_label': grade_label, 'grade_color': grade_color,
        'health_score': health_score, 'coverage_pct': coverage_pct,
        'blended_score': round(blended, 1),
        'open_dev_n': open_dev_n, 'inspected': inspected,
        'never_synced': never_synced,
        'trend': trend, 'overall_delta': overall_delta,
        'trend_max': max((t['p0'] + t['p1'] + t['p2'] + 1) for t in trend) if trend else 1,
        'trend_first_total': (trend[0]['p0'] + trend[0]['p1'] + trend[0]['p2']) if trend else 0,
        'trend_last_total': (trend[-1]['p0'] + trend[-1]['p1'] + trend[-1]['p2']) if trend else 0,
        'failure_modes': failure_modes[:12],
        'fm_max': failure_modes[0]['count'] if failure_modes else 1,
        'risk_by_role': risk_by_role, 'risk_by_site': risk_by_site,
        'recommendations': recommendations, 'narrative': narrative,
    }


def fleet_report(request):
    """整体巡检报告：项目设备整体态势（独立于单次巡检）。"""
    site = request.GET.get('site', '')
    stats = _fleet_stats(site)

    onsite = stats['onsite_total'] or 1
    C = 289
    ok_len = round(stats['synced_ok'] / onsite * C)
    warn_len = round(stats['warn_count'] / onsite * C)
    sync_len = max(C - ok_len - warn_len, 0)
    warn_offset = ok_len + warn_len

    return render(request, 'fleet_report.html', {
        'stats': stats,
        'site': site,
        'sites': _cmdb_sites(),
        'donut': {'ok': ok_len, 'warn': warn_len, 'sync': sync_len,
                  'warn_offset': warn_offset, 'C': C},
        'now': datetime.datetime.now().strftime('%Y-%m-%d %H:%M'),
    })


def fleet_report_export(request):
    """整体巡检报告导出为 Excel。"""
    from django.http import HttpResponse
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    site = request.GET.get('site', '')
    s = _fleet_stats(site)
    wb = openpyxl.Workbook()
    head_fill = PatternFill('solid', fgColor='0F6E56')
    head_font = Font(color='FFFFFF', bold=True)
    title_font = Font(size=14, bold=True)

    def style_head(ws, row, n):
        for c in range(1, n + 1):
            ws.cell(row=row, column=c).fill = head_fill
            ws.cell(row=row, column=c).font = head_font

    # ── 整体概览 ──
    ws = wb.active
    ws.title = '整体概览'
    ws.append(['整体设备态势报告', site or '全部站点',
               datetime.datetime.now().strftime('%Y-%m-%d %H:%M')])
    ws['A1'].font = title_font
    ws.append([])
    overview = [
        ('设备总数', s['total']),
        ('已建台账(本站点)', s['cmdb_count']),
        ('资源预警设备', s['warn_count']),
        ('待处理异常(累计未确认)', s['total_anom']),
        ('高危项 P0+P1', s['high']),
        ('  - P0', s['p0']),
        ('  - P1', s['p1']),
        ('  - P2', s['p2']),
        ('资产健康率(%)', s['healthy_pct']),
        ('未同步台账设备', s['not_synced']),
        ('最近巡检时间', s['latest_run_time']),
        ('最近台账同步', str(s['latest_sync']) if s['latest_sync'] else '无'),
    ]
    ws.append(['指标', '数值'])
    style_head(ws, 3, 2)
    for k, v in overview:
        ws.append([k, v])

    # ── 站点 / 角色分布 ──
    ws2 = wb.create_sheet('分布')
    ws2.append(['维度', '名称', '设备数'])
    style_head(ws2, 1, 3)
    for d in s['site_dist']:
        ws2.append(['站点', d['name'], d['count']])
    for d in s['role_dist']:
        ws2.append(['角色', d['name'], d['count']])
    for d in s['site_anom_dist']:
        ws2.append(['异常-站点', d['name'], d['count']])
    for d in s['role_anom_dist']:
        ws2.append(['异常-角色', d['name'], d['count']])

    # ── 资源水位预警 ──
    ws3 = wb.create_sheet('资源预警')
    ws3.append(['设备', '站点', '角色', 'CPU(5s)%', '内存空闲%', '运行天数', '预警项', '型号', '序列号'])
    style_head(ws3, 1, 9)
    for w in s['warn_list']:
        ws3.append([w['name'], w['site'], w['role'], w['cpu'], w['mem'],
                    w['uptime'], w['reasons'], w['model'], w['serial']])

    # ── 问题清单 ──
    ws4 = wb.create_sheet('问题清单')
    ws4.append(['等级', '巡检项', '异常说明', '影响设备数', '涉及设备'])
    style_head(ws4, 1, 5)
    for it in s['issues']:
        ws4.append([it['severity'], it['command'], it['notes'],
                    it['device_count'], ', '.join(it['devices'])])

    # ── 异常 Top 设备 ──
    ws5 = wb.create_sheet('异常Top设备')
    ws5.append(['设备', '异常项数', '最高级别', 'IP', '站点', '角色'])
    style_head(ws5, 1, 6)
    for d in s['top_devices']:
        m = d['meta'] or {}
        ws5.append([d['name'], d['count'], d['sev'],
                    m.get('ip', ''), m.get('site', ''), m.get('role', '')])

    # ── 最近巡检 ──
    ws6 = wb.create_sheet('最近巡检')
    ws6.append(['巡检时间', '操作人', '设备数', '正常', '异常', '失败', '结果'])
    style_head(ws6, 1, 7)
    for r in s['recent']:
        ws6.append([r['time'], r['operator'], r['device_count'], r['ok'],
                    r['anomaly'], r['failed'], r['result']])

    # ── 跨轮次趋势 ──
    ws7 = wb.create_sheet('异常趋势')
    ws7.append(['巡检轮次', '设备数', 'P0', 'P1', 'P2', '异常项合计', '异常率%', '较上轮变化'])
    style_head(ws7, 1, 8)
    for t in s['trend']:
        ws7.append([t['time'], t['device_count'], t['p0'], t['p1'], t['p2'],
                    t['p0'] + t['p1'] + t['p2'], t['rate'],
                    (f'+{t["delta"]}' if (t['delta'] or 0) > 0 else
                     (str(t['delta']) if (t['delta'] or 0) < 0 else '-')) if t['delta'] is not None else '-'])

    # ── 故障模式归因 ──
    ws8 = wb.create_sheet('故障模式')
    ws8.append(['巡检项', '等级', '影响设备数', '命中次数', '性质'])
    style_head(ws8, 1, 5)
    for fm in s['failure_modes']:
        ws8.append([fm['command'], fm['severity'], fm['device_count'],
                    fm['count'], '系统性' if fm['systemic'] else '局部'])

    resp = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = (
        f'attachment; filename=fleet_report_{site or "all"}_'
        f'{datetime.datetime.now():%Y%m%d%H%M}.xlsx')
    wb.save(resp)
    return resp


# ════════════════════════════════════════════════
# CMDB 台账查询（设备 / 接口 / 链路 / IP）
# ════════════════════════════════════════════════
def _cmdb_sites():
    from app02.models import CmdbDevice
    return list(CmdbDevice.objects.values_list('site', flat=True)
                .distinct().exclude(site='').order_by('site'))


def cmdb_device_list(request):
    """CMDB 设备台账：站点 + 健康状态 + 关键字筛选"""
    from app02.models import CmdbDevice
    from django.db.models import Q, Case, When, Value, BooleanField
    site = request.GET.get('site', '')
    status = request.GET.get('status', '')   # all / ok / warn
    q = request.GET.get('q', '').strip()
    qs = CmdbDevice.objects.all()
    if site:
        qs = qs.filter(site=site)
    if status == 'warn':
        qs = qs.filter(Q(cpu_5s__gte=80) | Q(mem_free_ratio__lt=20))
    elif status == 'ok':
        qs = qs.exclude(Q(cpu_5s__gte=80) | Q(mem_free_ratio__lt=20))
    if q:
        qs = qs.filter(Q(name__icontains=q) | Q(model__icontains=q) |
                       Q(mgmt_ip__icontains=q) | Q(serial__icontains=q))
    qs = qs.annotate(is_warn=Case(
        When(Q(cpu_5s__gte=80) | Q(mem_free_ratio__lt=20), then=Value(True)),
        default=Value(False), output_field=BooleanField()))
    return render(request, 'cmdb_device.html', {
        'queryset': qs, 'sites': _cmdb_sites(),
        'site': site, 'status': status, 'q': q,
    })


def cmdb_interface_list(request):
    """CMDB 接口台账：站点 + 状态(UP/DOWN) + 关键字筛选"""
    from app02.models import CmdbInterface
    from django.db.models import Q
    site = request.GET.get('site', '')
    status = request.GET.get('status', '')   # all / up / down
    q = request.GET.get('q', '').strip()
    qs = CmdbInterface.objects.select_related('device')
    if site:
        qs = qs.filter(device__site=site)
    if status == 'up':
        qs = qs.filter(oper_status='UP')
    elif status == 'down':
        qs = qs.filter(oper_status='DOWN')
    if q:
        qs = qs.filter(Q(name__icontains=q) | Q(device__name__icontains=q) |
                       Q(description__icontains=q))
    return render(request, 'cmdb_interface.html', {
        'queryset': qs, 'sites': _cmdb_sites(),
        'site': site, 'status': status, 'q': q,
    })


def cmdb_link_list(request):
    """CMDB 邻居链路(LLDP 拓扑)：站点 + 关键字筛选"""
    from app02.models import CmdbNeighborLink
    from django.db.models import Q
    site = request.GET.get('site', '')
    q = request.GET.get('q', '').strip()
    qs = CmdbNeighborLink.objects.select_related('device')
    if site:
        qs = qs.filter(device__site=site)
    if q:
        qs = qs.filter(Q(local_port__icontains=q) | Q(peer_device__icontains=q) |
                       Q(peer_port__icontains=q) | Q(device__name__icontains=q))
    return render(request, 'cmdb_link.html', {
        'queryset': qs, 'sites': _cmdb_sites(), 'site': site, 'q': q,
    })


def cmdb_ip_list(request):
    """CMDB 接口IP：站点 + 关键字筛选"""
    from app02.models import CmdbIpSubnet
    from django.db.models import Q
    site = request.GET.get('site', '')
    q = request.GET.get('q', '').strip()
    qs = CmdbIpSubnet.objects.select_related('device')
    if site:
        qs = qs.filter(device__site=site)
    if q:
        qs = qs.filter(Q(cidr__icontains=q) | Q(interface_name__icontains=q) |
                       Q(vrf__icontains=q) | Q(device__name__icontains=q))
    return render(request, 'cmdb_ip.html', {
        'queryset': qs, 'sites': _cmdb_sites(), 'site': site, 'q': q,
    })
