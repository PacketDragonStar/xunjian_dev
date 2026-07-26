# -*- coding: utf-8 -*-
"""导出全部巡检项到 Excel，供人工审阅/修改，再用 import_checkitems_excel 回灌。

用法:
    python manage.py export_checkitems_excel
    python manage.py export_checkitems_excel --out 巡检项配置_2026.xlsx
"""
import json
import os

from django.core.management.base import BaseCommand
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app02.models import CheckItem

# 列定义: (字段key, 表头, 是否可编辑, 备注)
COLUMNS = [
    ('command',        '命令',            False, '唯一标识，请勿修改'),
    ('name',           '名称',            True,  '巡检项显示名称'),
    ('enabled',        '是否启用',        True,  'Y / N（或 是/否/1/0/true/false）'),
    ('parser',         '解析器',          True,  'raw / regex / strip_ts / textfsm'),
    ('parser_config',  '解析器配置JSON',  True,  'JSON，可空。如 {"pattern":"..."}'),
    ('checker',        '检查器',          True,  'baseline / threshold / count / contains / custom'),
    ('checker_config', '检查器配置JSON',  True,  'JSON，可空。如 {"similarity":1.0} 或 {"warning":75}'),
    ('severity',       '严重级别',        True,  'P0 / P1 / P2'),
    ('timeout',        '超时秒',          True,  '整数，如 30'),
    ('error_note',     '异常提示',        True,  '判定异常时展示的文字'),
    ('fix_suggestion', '整改建议',        True,  '可空'),
    ('extract_parser', '字段提取器',      True,  '空 / memory'),
    ('compare_strip',  '对比清洗配置JSON', True,  'JSON，可空。如 {"head_lines":3,"skip_patterns":["^2026-"]}'),
    ('device_groups',  '挂载分组',        False, '只读参考：当前挂载到的设备分组'),
    ('当前状态',        '当前状态',        False, '只读参考：由 enabled+checker 自动计算'),
]

PARSER_CHOICES = ['raw', 'regex', 'strip_ts', 'textfsm']
CHECKER_CHOICES = ['baseline', 'threshold', 'count', 'contains', 'custom']
EXTRACT_CHOICES = ['', 'memory']
SEVERITY_CHOICES = ['P0', 'P1', 'P2']

STATUS_MAP = {
    'raw':      '仅采集(不判定)',
    'contains': '仅采集(不判定)',
    'baseline': '基线对比',
    'custom':   '自定义函数',
    'threshold':'阈值判断',
    'count':    '关键字计数',
}


def _json_dump(v):
    if v is None:
        return ''
    # 库里有时会把 JSON 以「字符串」形式存（如 '{"func":...}'），
    # 这里先尝试解析成对象再 dump，保证 Excel 内是规整的 JSON。
    if isinstance(v, str):
        s = v.strip()
        if s and s[0] in '{["':
            try:
                v = json.loads(s)
            except (ValueError, TypeError):
                return s
        else:
            return s
    return json.dumps(v, ensure_ascii=False)


def _status_label(ci):
    if not ci.enabled:
        return '未启用'
    return STATUS_MAP.get(ci.checker, ci.checker)


class Command(BaseCommand):
    help = '导出全部巡检项为 Excel（含现状说明），便于人工批量修改'

    def add_arguments(self, parser):
        parser.add_argument('--out', default=None,
                            help='输出 xlsx 路径（默认：项目根 巡检项配置_导出.xlsx）')

    def handle(self, *args, **opts):
        out = opts['out']
        if not out:
            base = os.path.dirname(os.path.dirname(os.path.dirname(
                os.path.dirname(os.path.abspath(__file__)))))
            out = os.path.join(base, '巡检项配置_导出.xlsx')

        items = list(CheckItem.objects.all().order_by('command'))

        wb = Workbook()
        ws = wb.active
        ws.title = '巡检项'

        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill('solid', fgColor='0F6E56')  # 青蓝主题
        center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left = Alignment(horizontal='left', vertical='center', wrap_text=True)
        thin = Side(style='thin', color='D0D0D0')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        # 表头
        for c, (key, title, editable, note) in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=1, column=c, value=title)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = border
            if not editable:
                cell.comment = None  # 备注见说明 sheet

        # 数据行
        for r, ci in enumerate(items, start=2):
            groups = ', '.join(g.name for g in ci.groups.all())
            row_vals = {
                'command': ci.command,
                'name': ci.name,
                'enabled': 'Y' if ci.enabled else 'N',
                'parser': ci.parser,
                'parser_config': _json_dump(ci.parser_config),
                'checker': ci.checker,
                'checker_config': _json_dump(ci.checker_config),
                'severity': ci.severity,
                'timeout': ci.timeout,
                'error_note': ci.error_note or '',
                'fix_suggestion': ci.fix_suggestion or '',
                'extract_parser': ci.extract_parser or '',
                'compare_strip': _json_dump(ci.compare_strip),
                'device_groups': groups,
                '当前状态': _status_label(ci),
            }
            for c, (key, title, editable, note) in enumerate(COLUMNS, start=1):
                cell = ws.cell(row=r, column=c, value=row_vals.get(key, ''))
                cell.border = border
                # 只读列浅灰底色
                if not editable:
                    cell.fill = PatternFill('solid', fgColor='F2F2F2')
                    cell.alignment = left
                else:
                    cell.alignment = left
                # 高亮异常状态行
                if key == '当前状态':
                    if row_vals['当前状态'] in ('未启用', '仅采集(不判定)'):
                        cell.font = Font(color='C0392B', bold=True)
                    else:
                        cell.font = Font(color='0F6E56', bold=True)

        # 列宽
        widths = [34, 20, 10, 12, 30, 12, 30, 10, 8, 22, 22, 12, 34, 26, 16]
        for c, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(c)].width = w

        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = f'A1:{get_column_letter(len(COLUMNS))}{len(items)+1}'

        # ---- 说明 sheet ----
        ws2 = wb.create_sheet('说明')
        lines = [
            ('巡检项 Excel 使用说明', True),
            ('', False),
            ('1. 主表「巡检项」每一行是一个巡检项，「命令」列是唯一标识，请勿修改。', False),
            ('2. 灰色列（命令 / 挂载分组 / 当前状态）为只读参考，导入时忽略，不要填。', False),
            ('3. 你只需修改白色可编辑列，改完把文件发回，我用 import_checkitems_excel 回灌。', False),
            ('', False),
            ('各可编辑列取值说明：', True),
            ('  • 是否启用: Y / N（也接受 是/否/1/0/true/false；留空=保持原值）', False),
            ('  • 解析器 parser 合法值: ' + ' / '.join(PARSER_CHOICES), False),
            ('  • 检查器 checker 合法值: ' + ' / '.join(CHECKER_CHOICES), False),
            ('  • 字段提取器 extract_parser 合法值: ' + (' / '.join(EXTRACT_CHOICES) or '空'), False),
            ('  • 严重级别 severity 合法值: ' + ' / '.join(SEVERITY_CHOICES), False),
            ('  • 超时秒 timeout: 整数，如 30', False),
            ('  • 带 JSON 的列（解析器配置/检查器配置/对比清洗配置）必须写合法 JSON，可空留白。', False),
            ('     例: 检查器配置={"similarity":1.0}  对比清洗配置={"head_lines":3,"skip_patterns":["^2026-"]}', False),
            ('', False),
            ('常见检查器含义：', True),
            ('  • baseline 基线文本对比: 与「基线巡检」的该命令输出逐行 diff，变了就告警（需先设一条基线记录）', False),
            ('  • threshold 阈值判断: 由检查器配置里的阈值或自定义函数判定', False),
            ('  • count 关键字计数: 统计匹配关键字的条数', False),
            ('  • contains 包含检查: 仅采集/简单包含判断（常用于纯采集项）', False),
            ('  • custom 自定义函数: 调用 custom_checks.py 里 @register_checker 的函数（如 check_flash_usage / check_logbuffer）', False),
            ('', False),
            ('「当前状态」列自动计算，帮你看清现状：', True),
            ('  • 未启用: enabled=N，巡检不会执行该项', False),
            ('  • 仅采集(不判定): checker=raw 或 contains，只存回显、不告警', False),
            ('  • 基线对比 / 自定义函数 / 阈值判断 / 关键字计数: 会实际参与判定', False),
            ('', False),
            ('回灌命令（改完告诉我，我来跑）：', True),
            ('  python manage.py import_checkitems_excel --in <你改好的文件.xlsx>', False),
            ('  （加 --dry-run 可先预览将要变更的内容，不实际写入）', False),
        ]
        for r, (text, bold) in enumerate(lines, start=1):
            cell = ws2.cell(row=r, column=1, value=text)
            cell.font = Font(bold=bold, size=12 if bold else 11,
                             color='0F6E56' if (bold and r == 1) else '000000')
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        ws2.column_dimensions['A'].width = 110

        wb.save(out)
        self.stdout.write(self.style.SUCCESS(
            f'已导出 {len(items)} 个巡检项 -> {out}'))
