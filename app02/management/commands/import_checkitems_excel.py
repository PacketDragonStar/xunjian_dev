# -*- coding: utf-8 -*-
"""按 Excel 回灌巡检项配置（配合 export_checkitems_excel 使用）。

用法:
    python manage.py import_checkitems_excel --in 巡检项配置_导出.xlsx
    python manage.py import_checkitems_excel --in 巡检项配置_导出.xlsx --dry-run

说明:
    - 以「命令」列作为唯一匹配键，找不到命令的行会被跳过并提示。
    - 只读列（命令 / 挂载分组 / 当前状态）忽略。
    - 仅更新可编辑字段；空白的 JSON 列视为 None（清空）；空白的「是否启用」保持原值。
"""
import json

from django.core.management.base import BaseCommand, CommandError
from openpyxl import load_workbook

from app02.models import CheckItem

# (列名 -> 模型字段) 可编辑映射
EDIT_MAP = {
    '名称':           'name',
    '是否启用':       'enabled',
    '解析器':         'parser',
    '解析器配置JSON': 'parser_config',
    '检查器':         'checker',
    '检查器配置JSON': 'checker_config',
    '严重级别':       'severity',
    '超时秒':         'timeout',
    '异常提示':       'error_note',
    '整改建议':       'fix_suggestion',
    '字段提取器':     'extract_parser',
    '对比清洗配置JSON':'compare_strip',
}
JSON_FIELDS = {'parser_config', 'checker_config', 'compare_strip'}
BOOL_TRUE = {'y', '是', '1', 'true', '启用', 'yes'}
BOOL_FALSE = {'n', '否', '0', 'false', '停用', 'no'}


def _parse_json_cell(val):
    """空白 -> None；可解析 JSON -> 解析后对象；否则保留原始字符串。"""
    if val is None:
        return None
    s = str(val).strip()
    if s == '':
        return None
    try:
        return json.loads(s)
    except (ValueError, TypeError):
        return s  # 非 JSON，保留原字符串（如纯文本配置）


def _coerce(field, val):
    if field == 'enabled':
        if val is None or str(val).strip() == '':
            return None  # 保持原值
        low = str(val).strip().lower()
        if low in BOOL_TRUE:
            return True
        if low in BOOL_FALSE:
            return False
        raise CommandError(f'无法识别的启用值: {val!r}（应为 Y/N）')
    if field == 'timeout':
        return int(str(val).strip())
    if field in JSON_FIELDS:
        return _parse_json_cell(val)
    if field == 'fix_suggestion' and (val is None or str(val).strip() == ''):
        return None  # 空整改建议等价于未填（模型允许 null）
    return str(val).strip() if val is not None else ''


class Command(BaseCommand):
    help = '按 Excel 回灌巡检项配置'

    def add_arguments(self, parser):
        parser.add_argument('--in', dest='infile', required=True,
                            help='Excel 路径（由 export_checkitems_excel 生成/修改）')
        parser.add_argument('--dry-run', action='store_true',
                            help='只预览将要做的变更，不实际写入数据库')

    def handle(self, *args, **opts):
        infile = opts['infile']
        dry = opts['dry_run']

        try:
            wb = load_workbook(infile, data_only=True)
        except Exception as e:
            raise CommandError(f'读取 Excel 失败: {e}')

        if '巡检项' not in wb.sheetnames:
            raise CommandError('找不到「巡检项」工作表，请确认是 export_checkitems_excel 生成的文件')

        ws = wb['巡检项']
        header = [c.value for c in ws[1]]
        # 列名 -> 列索引
        col_idx = {name: i for i, name in enumerate(header) if name}

        missing = [k for k in EDIT_MAP if k not in col_idx]
        if missing:
            raise CommandError(f'Excel 缺少必要列: {missing}')

        updated, skipped, notfound = 0, 0, []
        for r in range(2, ws.max_row + 1):
            row = [ws.cell(row=r, column=i + 1).value for i in range(len(header))]
            cmd = (row[col_idx['命令']] or '').strip() if col_idx.get('命令') is not None else ''
            if not cmd:
                continue
            try:
                ci = CheckItem.objects.get(command=cmd)
            except CheckItem.DoesNotExist:
                notfound.append(cmd)
                continue

            changes = {}
            for col_name, field in EDIT_MAP.items():
                raw = row[col_idx[col_name]]
                if field == 'enabled' and (raw is None or str(raw).strip() == ''):
                    # 空白启用列 -> 保持原值，跳过
                    continue
                try:
                    new_val = _coerce(field, raw)
                except CommandError as e:
                    raise CommandError(f'第 {r} 行 [{cmd}] {col_name} 列: {e}')
                old_val = getattr(ci, field)
                # 比较（JSON 字段按内容比较）
                if new_val != old_val:
                    changes[field] = (old_val, new_val)

            if not changes:
                skipped += 1
                continue

            if dry:
                self.stdout.write(f'[预览] {cmd}:')
                for f, (o, n) in changes.items():
                    self.stdout.write(f'    {f}: {o!r} -> {n!r}')
            else:
                for f, (o, n) in changes.items():
                    setattr(ci, f, n)
                ci.save()
            updated += 1

        self.stdout.write('-' * 50)
        self.stdout.write(f'匹配成功并应用变更: {updated}')
        self.stdout.write(f'无变化跳过:         {skipped}')
        if notfound:
            self.stdout.write(f'未找到命令(跳过):   {len(notfound)} -> {notfound}')
        if dry:
            self.stdout.write(self.style.WARNING('（--dry-run 预览模式，未写入数据库）'))
        else:
            self.stdout.write(self.style.SUCCESS('回灌完成。runserver 已热重载，下次巡检生效。'))
